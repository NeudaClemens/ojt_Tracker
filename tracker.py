import csv
import os
import json
import re
import calendar as pycalendar
import secrets
import hashlib
import hmac
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog, ttk
from openpyxl import Workbook, load_workbook

try:
    import cv2
except Exception:
    cv2 = None
try:
    import qrcode
except Exception:
    qrcode = None
try:
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas as pdf_canvas
except Exception:
    pdf_canvas = None

DATE_FMT = "%Y-%m-%d"
TIME_FMT_DISPLAY = "%I:%M %p"

TIME_IN_GRACE_MINUTES = 5
SIGN_OUT_ALLOWED_FROM_HOUR = 17  # 5:00 PM
LATE_AFTER_TIME = "09:00 AM"     # automatic late after 9:00 AM

DATA_DIR = Path.home() / "Documents" / "OJT_Tracker"
XLSX_FILE = DATA_DIR / "ojt_tracker_records.xlsx"
CSV_FILE = DATA_DIR / "ojt_tracker_records.csv"
ADMIN_CONFIG_FILE = DATA_DIR / "admin_config.json"
QR_DIR = DATA_DIR / "qr_codes"
DEFAULT_ADMIN_PASSWORD = os.environ.get("OJT_ADMIN_PASSWORD", "admin123")
USER_PROFILE_FIELDS = ("first_name", "last_name", "job_role",
                       "student_id", "course", "university", "password")

# Global design system
PRIMARY = "#1F4E79"  # Deep Blue
ACCENT = "#2F80ED"  # Bright Blue
SUCCESS = "#27AE60"
WARNING = "#F2C94C"
DANGER = "#EB5757"
BG = "#F5F7FA"
CARD = "#FFFFFF"
TEXT_DARK = "#1A1A1A"
TEXT_LIGHT = "#6B7280"
BORDER = "#E0E6ED"

FONT_TITLE = ("Segoe UI", 26, "bold")
FONT_HEADER = ("Segoe UI", 16, "bold")
FONT_NORMAL = ("Segoe UI", 11)
FONT_SMALL = ("Segoe UI", 9)
FONT_NUMBER = ("Segoe UI", 22, "bold")
FONT_BTN_SMALL = ("Segoe UI", 10)


@dataclass
class AttendanceEntry:
    work_date: date
    time_in: datetime
    time_out: datetime
    hours: float


# ===== Password hashing helpers (PBKDF2-HMAC-SHA256) =====
_PW_PREFIX = "pbkdf2_sha256"
_PW_ITERS = 200_000


def _hash_password(plain: str) -> str:
    plain = (plain or "").strip()
    if not plain:
        return ""
    salt = secrets.token_hex(16)
    dk = hashlib.pbkdf2_hmac("sha256", plain.encode("utf-8"),
                             salt.encode("utf-8"), _PW_ITERS)
    return f"{_PW_PREFIX}${_PW_ITERS}${salt}${dk.hex()}"


def _is_hashed_password(value: str) -> bool:
    return isinstance(value, str) and value.startswith(f"{_PW_PREFIX}$")


def _verify_password(plain: str, stored: str) -> bool:
    plain = (plain or "").strip()
    stored = (stored or "").strip()
    if not stored or not plain:
        return False

    # Legacy plaintext fallback: allow login, then we can upgrade on save/change.
    if not _is_hashed_password(stored):
        return hmac.compare_digest(plain, stored)

    try:
        _prefix, iters_s, salt, hexhash = stored.split("$", 3)
        iters = int(iters_s)
        dk = hashlib.pbkdf2_hmac(
            "sha256",
            plain.encode("utf-8"),
            salt.encode("utf-8"),
            iters,
        )
        return hmac.compare_digest(dk.hex(), hexhash)
    except Exception:
        return False


class OJTTrackerApp:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("AUTOLIV")
        self.root.state("zoomed")
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)
        DATA_DIR.mkdir(parents=True, exist_ok=True)

        self.users_data = {}
        self._last_active_user = ""
        self.is_admin = False
        self.admin_password = self._load_admin_password()

        self.active_user_var = tk.StringVar(value="")
        self.active_user_display_var = tk.StringVar(value="")
        self.required_hours_var = tk.StringVar(value="500")
        self.date_var = tk.StringVar(value=date.today().strftime(DATE_FMT))
        self.time_in_var = tk.StringVar()
        self.time_out_var = tk.StringVar()
        self.progress_value_var = tk.DoubleVar(value=0.0)
        self.progress_percent_var = tk.StringVar(value="0.00%")
        self.status_var = tk.StringVar(value="Status: Just Started")
        self.overview_date_var = tk.StringVar(value="")
        self.overview_time_var = tk.StringVar(value="")
        self.overview_completion_var = tk.StringVar(value="0.00%")
        self.scan_feedback_var = tk.StringVar(value="Ready to scan QR")
        self.date_filter_var = tk.StringVar(value="")
        self.time_in_filter_var = tk.StringVar(value="")
        self.time_out_filter_var = tk.StringVar(value="")
        self.from_date_var = tk.StringVar(value="")
        self.to_date_var = tk.StringVar(value="")
        self.summary_var = tk.StringVar(
            value="Total Days: 0 | Avg Hours: 0.00 | Late Count: 0")
        self.intern_count_var = tk.StringVar(value="Interns: 0")

        self.completed_hours_var = tk.StringVar(value="0.00 hours")
        self.remaining_hours_var = tk.StringVar(value="0.00")
        self.required_hours_display_var = tk.StringVar(value="500")
        self.profile_name_var = tk.StringVar(value="-")
        self.profile_job_role_var = tk.StringVar(value="-")
        self.profile_required_hours_var = tk.StringVar(value="500")
        self.profile_student_id_var = tk.StringVar(value="-")
        self.profile_course_var = tk.StringVar(value="-")
        self.profile_university_var = tk.StringVar(value="-")
        self._display_to_user = {}
        self._user_to_display = {}
        self.user_profile_window = None
        self.user_profile_widgets = {}
        self.main_ui_window = None
        self.main_ui_widgets = {}
        # Keep user profiles in a separate toplevel window (NOT a second Tk()).
        self.user_profile_separate = True

        # When filters are active, we keep full records intact and just show a filtered view.
        self._filters_active = False

        self._build_ui()
        self._load_records()
        self._ensure_default_user()
        self._load_active_user_into_view()
        self._heartbeat()
        self.root.after(50, self._start_main_ui)

    def _effective_admin_button_text(self):
        return "Admin Log Out" if self.is_admin else "Admin Login"

    def _admin_button_action(self):
        """Called by the top admin header button — logs out when already admin, otherwise opens login."""
        if self.is_admin:
            self._admin_logout()
        else:
            self._toggle_admin()

    def _main_admin_button_text(self):
        return "Open Admin UI" if self.is_admin else "Admin Login"

    def _start_main_ui(self):
        self._open_main_ui_window()
        self.root.withdraw()

    def _center_toplevel(self, win):
        win.update_idletasks()
        width = win.winfo_width()
        height = win.winfo_height()
        screen_w = win.winfo_screenwidth()
        screen_h = win.winfo_screenheight()
        pos_x = max((screen_w - width) // 2, 0)
        pos_y = max((screen_h - height) // 2, 0)
        win.geometry(f"+{pos_x}+{pos_y}")

    # Color helpers
    def _hex_to_rgb(self, hex_color: str):
        hex_color = hex_color.lstrip("#")
        return tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))

    def _rgb_to_hex(self, rgb):
        return "#%02x%02x%02x" % rgb

    # Animated progress for top-level progressbar
    def _animate_progress(self, target: float, duration: int = 400):
        try:
            start = float(self.progress_value_var.get())
        except Exception:
            start = 0.0
        steps = max(int(duration / 30), 1)
        delta = (target - start) / steps

        def step(i=0, value=start):
            if i >= steps:
                self.progress_value_var.set(target)
                self.progress_percent_var.set(f"{target:.2f}%")
                return
            value += delta
            self.progress_value_var.set(value)
            self.progress_percent_var.set(f"{value:.2f}%")
            self.root.after(int(duration / steps), lambda: step(i + 1, value))

        step()

    # Animated canvas progress (user profile)
    def _animate_canvas_progress(self, canvas: tk.Canvas, target_percent: float, color: str, duration: int = 400):
        canvas.update_idletasks()
        width = canvas.winfo_width()
        height = canvas.winfo_height() or 20
        if not width or width <= 1:
            canvas.after(50, lambda: self._animate_canvas_progress(
                canvas, target_percent, color, duration))
            return

        try:
            start = float(getattr(canvas, "_last_percent", 0.0))
        except Exception:
            start = 0.0

        steps = max(int(duration / 30), 1)
        delta = (target_percent - start) / steps

        canvas.delete("bg")
        canvas.create_rectangle(
            0, 0, width, height, fill="#E8EAED", outline="#E8EAED", tags=("bg",))

        def step(i=0, p=start):
            if i >= steps:
                p = target_percent
            else:
                p += delta

            canvas.delete("fill")
            fill_w = int(width * (max(0.0, min(100.0, p)) / 100.0))
            if fill_w > 0:
                canvas.create_rectangle(
                    0, 0, fill_w, height, fill=color, outline=color, tags=("fill",))

            canvas.delete("border")
            canvas.create_rectangle(
                0, 0, width, height, outline="#AAB0B6", tags=("border",))
            canvas._last_percent = p
            if i < steps:
                canvas.after(int(duration / steps), lambda: step(i + 1, p))

        step()

    def _new_user_payload(self, required_hours="500", profile=None):
        profile = profile or {}
        pw = (profile.get("password", "") or "").strip()
        pw_store = _hash_password(
            pw) if pw and not _is_hashed_password(pw) else pw
        return {
            "required_hours": str(required_hours or "500"),
            "records": [],
            "first_name": profile.get("first_name", "").strip(),
            "last_name": profile.get("last_name", "").strip(),
            "job_role": profile.get("job_role", "").strip(),
            "student_id": profile.get("student_id", "").strip(),
            "course": profile.get("course", "").strip(),
            "university": profile.get("university", "").strip(),
            "password": pw_store,
            "is_deleted": False,
        }

    def _normalize_user_payload(self, user_key: str, payload):
        payload = dict(payload or {})
        normalized = self._new_user_payload(
            required_hours=payload.get("required_hours", "500"), profile=payload)
        normalized["records"] = payload.get("records", []) or []
        normalized["is_deleted"] = str(payload.get(
            "is_deleted", "0")).strip().lower() in ("1", "true", "yes")

        # Keep stored password as-is (may be legacy plaintext); we verify accordingly.
        normalized["password"] = str(payload.get("password", "") or "").strip()

        if not normalized["first_name"] and not normalized["last_name"]:
            parts = user_key.split(" ", 1)
            normalized["first_name"] = parts[0] if parts else user_key
            normalized["last_name"] = parts[1] if len(parts) > 1 else ""
        return normalized

    def _full_name_for_user(self, user_key: str):
        payload = self.users_data.get(user_key, {})
        first = str(payload.get("first_name", "")).strip()
        last = str(payload.get("last_name", "")).strip()
        full = f"{first} {last}".strip()
        return full or user_key

    def _display_label_for_user(self, user_key: str):
        payload = self.users_data.get(user_key, {})
        full_name = self._full_name_for_user(user_key)
        student_id = str(payload.get("student_id", "")).strip()
        if student_id:
            return f"{full_name} ({student_id})"
        return full_name

    def _active_user_keys(self):
        return [k for k, v in self.users_data.items() if not v.get("is_deleted")]

    def _is_auto_default_user(self, key: str, payload):
        if key != "Default":
            return False
        data = payload or {}
        if data.get("records"):
            return False
        return (
            str(data.get("first_name", "")).strip() == "Default"
            and not str(data.get("last_name", "")).strip()
            and not str(data.get("job_role", "")).strip()
            and not str(data.get("student_id", "")).strip()
            and not str(data.get("course", "")).strip()
            and not str(data.get("university", "")).strip()
            and not str(data.get("password", "")).strip()
        )

    def _build_ui(self):
        self.bg_color = BG
        self.card_bg = CARD
        self.primary_blue = PRIMARY
        self.secondary_blue = ACCENT
        self.text_dark = TEXT_DARK
        self.text_light = TEXT_LIGHT
        self.border_color = BORDER
        self.accent_green = SUCCESS
        self.accent_yellow = WARNING
        self.accent_red = DANGER

        self.root.configure(bg=BG)
        self.style = ttk.Style(self.root)
        self.style.theme_use("clam")

        self.style.configure(
            "Modern.Horizontal.TProgressbar",
            thickness=18,
            troughcolor="#E5EAF2",
            background=ACCENT,
            bordercolor=ACCENT,
        )

        self.style.configure(
            "Treeview",
            rowheight=35,
            font=FONT_NORMAL,
            background=self.card_bg,
            fieldbackground=self.card_bg,
            borderwidth=0,
        )
        self.style.map("Treeview", background=[
                       ("selected", PRIMARY)], foreground=[("selected", "white")])
        self.style.configure(
            "Treeview.Heading",
            font=FONT_NORMAL,
            background="#F8F9FA",
            foreground=TEXT_DARK,
            relief="flat",
        )
        self.style.map("Treeview.Heading", background=[("active", "#E8EAED")])

        self.style.configure(
            "Action.TButton",
            font=FONT_BTN_SMALL,
            padding=(6, 4),
            background="#D1D5DB",
            foreground=self.text_dark,
        )
        self.style.map("Action.TButton", background=[("active", "#BFC7CE")])

        self.style.configure(
            "ActionHover.TButton",
            font=FONT_BTN_SMALL,
            padding=(6, 4),
            background="#BFC7CE",
            foreground=self.text_dark,
        )

        self.style.configure(
            "ActionSmall.TButton",
            font=FONT_BTN_SMALL,
            padding=(6, 4),
            background="#D1D5DB",
            foreground=self.text_dark,
        )
        self.style.configure(
            "ActionSmallHover.TButton",
            font=FONT_BTN_SMALL,
            padding=(6, 4),
            background="#BFC7CE",
            foreground=self.text_dark,
        )
        self.style.configure(
            "DangerSmall.TButton",
            font=FONT_BTN_SMALL,
            padding=(6, 4),
            background="#D1D5DB",
            foreground=self.text_dark,
        )
        self.style.configure(
            "DangerSmallHover.TButton",
            font=FONT_BTN_SMALL,
            padding=(6, 4),
            background="#BFC7CE",
            foreground=self.text_dark,
        )
        self.style.configure(
            "SuccessSmall.TButton",
            font=FONT_BTN_SMALL,
            padding=(6, 4),
            background="#D1D5DB",
            foreground=self.text_dark,
        )
        self.style.configure(
            "SuccessSmallHover.TButton",
            font=FONT_BTN_SMALL,
            padding=(6, 4),
            background="#BFC7CE",
            foreground=self.text_dark,
        )
        self.style.configure(
            "OutlineSmall.TButton",
            font=FONT_BTN_SMALL,
            padding=(6, 4),
            background="#D1D5DB",
            foreground=self.text_dark,
            relief="flat",
            borderwidth=1,
        )
        self.style.configure(
            "OutlineSmallHover.TButton",
            font=FONT_BTN_SMALL,
            padding=(6, 4),
            background="#BFC7CE",
            foreground=self.text_dark,
        )

        self.style.configure(
            "Success.TButton",
            font=FONT_BTN_SMALL,
            padding=(6, 4),
            background="#D1D5DB",
            foreground=self.text_dark,
        )
        self.style.configure(
            "Danger.TButton",
            font=FONT_BTN_SMALL,
            padding=(6, 4),
            background="#D1D5DB",
            foreground=self.text_dark,
        )
        self.style.configure(
            "Outline.TButton",
            font=FONT_BTN_SMALL,
            padding=(6, 4),
            background="#D1D5DB",
            foreground=self.text_dark,
            relief="flat",
            borderwidth=1,
        )
        self.style.configure(
            "SuccessHover.TButton",
            font=FONT_BTN_SMALL,
            padding=(6, 4),
            background="#BFC7CE",
            foreground=self.text_dark,
        )
        self.style.configure(
            "DangerHover.TButton",
            font=FONT_BTN_SMALL,
            padding=(6, 4),
            background="#BFC7CE",
            foreground=self.text_dark,
        )
        self.style.configure(
            "OutlineHover.TButton",
            font=FONT_BTN_SMALL,
            padding=(6, 4),
            background="#BFC7CE",
            foreground=self.text_dark,
        )

        self.style.configure(
            "Secondary.TButton",
            font=FONT_BTN_SMALL,
            padding=(6, 4),
            background="#D1D5DB",
            foreground=self.text_dark,
        )
        self.style.configure(
            "SecondaryHover.TButton",
            font=FONT_BTN_SMALL,
            padding=(6, 4),
            background="#BFC7CE",
            foreground=self.text_dark,
        )

        self.style.configure(
            "TButton",
            font=FONT_BTN_SMALL,
            padding=(6, 4),
            background="#D1D5DB",
            foreground=self.text_dark,
        )

        self.style.configure(
            "GraySmall.TButton",
            font=FONT_BTN_SMALL,
            padding=(6, 4),
            background="#D1D5DB",
            foreground=self.text_dark,
            relief="flat",
            borderwidth=1,
        )
        self.style.configure(
            "GraySmallHover.TButton",
            font=FONT_BTN_SMALL,
            padding=(6, 4),
            background="#BFC7CE",
            foreground=self.text_dark,
        )

        container = tk.Frame(self.root, bg=BG, padx=20, pady=20)
        container.pack(fill="both", expand=True)

        header_row = tk.Frame(container, bg=BG)
        header_row.pack(fill="x", pady=(0, 15))

        title_container = tk.Frame(header_row, bg=BG)
        title_container.pack(side="left")

        tk.Label(title_container, text="AUTOLIV", font=FONT_TITLE,
                 bg=BG, fg=PRIMARY).pack(side="left")
        tk.Label(title_container, text=" | Internship Management System", font=FONT_HEADER,
                 bg=BG, fg=TEXT_LIGHT).pack(side="left", padx=(5, 0), pady=(4, 0))

        header_actions = tk.Frame(header_row, bg=BG)
        header_actions.pack(side="right")
        self.open_main_ui_btn = ttk.Button(
            header_actions, text="Back to Dashboard", command=self._open_main_ui_window, style="Secondary.TButton")
        self.open_main_ui_btn.pack(side="right", padx=(0, 10))
        self.open_main_ui_btn.bind("<Enter>", lambda e, s="SecondaryHover.TButton",
                                   w=self.open_main_ui_btn: self._on_btn_enter(e, hover_style=s, widget=w))
        self.open_main_ui_btn.bind("<Leave>", lambda e, s="Secondary.TButton",
                                   w=self.open_main_ui_btn: self._on_btn_leave(e, normal_style=s, widget=w))

        self.admin_btn = ttk.Button(header_actions, text="Admin Login", command=self._admin_button_action,
                                    style="Action.TButton")
        self.admin_btn.pack(side="right")
        self.admin_btn.bind("<Enter>", lambda e, s="ActionHover.TButton",
                            w=self.admin_btn: self._on_btn_enter(e, hover_style=s, widget=w))
        self.admin_btn.bind("<Leave>", lambda e, s="Action.TButton",
                            w=self.admin_btn: self._on_btn_leave(e, normal_style=s, widget=w))

        overview = tk.Frame(container, bg=BG)
        overview.pack(fill="x", pady=(0, 15))
        for var in [self.overview_date_var, self.overview_time_var]:
            tk.Label(overview, textvariable=var, bg=BG, fg=TEXT_LIGHT,
                     font=FONT_SMALL).pack(side="left", padx=(0, 15))

        status_dot = tk.Label(overview, text="🟢", bg=BG,
                              fg=SUCCESS, font=("Segoe UI", 12))
        status_dot.pack(side="left")
        tk.Label(overview, text="System Status: Online", bg=BG,
                 fg=TEXT_LIGHT, font=FONT_SMALL).pack(side="left", padx=(6, 0))

        self.admin_tools_frame = tk.Frame(container, bg=self.card_bg, bd=0, highlightthickness=1,
                                          highlightbackground=self.border_color)
        tk.Label(self.admin_tools_frame, text="Admin Controls", bg=self.card_bg, fg=self.text_dark,
                 font=("Segoe UI", 11, "bold")).grid(row=0, column=0, columnspan=4, sticky="w", padx=15, pady=(12, 10))
        button_specs = [
            ("Add User", self._add_user),
            ("Delete User Permanently", self._delete_user),
            ("Change Admin Password", self._change_admin_password),
            ("Export CSV", self._export_active_records_csv),
            ("Generate QR", self._generate_qr_for_active_user),
            ("Export PDF", self._export_active_records_pdf),
        ]
        style_map = {label: "GraySmall.TButton" for (
            label, _cb) in button_specs}
        hover_map = {"GraySmall.TButton": "GraySmallHover.TButton"}
        for idx, (label, callback) in enumerate(button_specs):
            st = style_map.get(label, "Secondary.TButton")
            btn = ttk.Button(self.admin_tools_frame,
                             text=label, command=callback, style=st)
            btn.grid(row=(idx // 4) + 1, column=idx %
                     4, padx=6, pady=(4, 8), sticky="ew")
            hover_style = hover_map.get(st)
            if hover_style:
                btn.bind("<Enter>", lambda e, s=hover_style,
                         w=btn: self._on_btn_enter(e, hover_style=s, widget=w))
                btn.bind("<Leave>", lambda e, n=st, w=btn: self._on_btn_leave(
                    e, normal_style=n, widget=w))
        for idx in range(4):
            self.admin_tools_frame.grid_columnconfigure(idx, weight=1)
        self.admin_tools_frame.pack_forget()

        def create_card(parent, heading, expand=False):
            card_outer = tk.Frame(parent, bg="#E9EEF6", padx=10, pady=8)
            card_outer.pack(fill="both" if expand else "x",
                            expand=expand, pady=8)

            card = tk.Frame(
                card_outer,
                bg=self.card_bg,
                bd=0,
                highlightthickness=1,
                highlightbackground=self.border_color,
                padx=0,
                pady=0,
            )
            card.pack(fill="both", expand=True, padx=4, pady=4)

            header = tk.Frame(card, bg="#F8F9FA")
            header.pack(fill="x")
            tk.Label(header, text=heading, font=("Segoe UI", 11, "bold"), bg="#F8F9FA", fg=self.text_dark).pack(
                anchor="w", padx=15, pady=10
            )

            sep = tk.Frame(card, height=1, bg=self.border_color)
            sep.pack(fill="x")

            body = tk.Frame(card, bg=self.card_bg, padx=20, pady=20)
            body.pack(fill="both", expand=True)
            return body

        self.columns_frame = tk.Frame(container, bg=BG)
        self.columns_frame.pack(fill="both", expand=True)
        self.columns_frame.grid_columnconfigure(0, weight=3, uniform="layout")
        self.columns_frame.grid_columnconfigure(1, weight=4, uniform="layout")
        self.columns_frame.grid_rowconfigure(0, weight=1)

        left_col = tk.Frame(self.columns_frame, bg=BG)
        left_col.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
        right_col = tk.Frame(self.columns_frame, bg=BG)
        right_col.grid(row=0, column=1, sticky="nsew", padx=(10, 0))

        user_frame = create_card(left_col, "Intern Profile")
        user_frame.columnconfigure(0, weight=0)
        user_frame.columnconfigure(1, weight=1)
        user_frame.columnconfigure(2, weight=0)
        tk.Label(user_frame, text="Intern Selection:", bg=self.card_bg, fg=self.text_light, font=("Segoe UI", 9)).grid(
            row=0, column=0, sticky="w")
        self.user_combo = ttk.Combobox(user_frame, textvariable=self.active_user_display_var, state="readonly",
                                       font=("Segoe UI", 10))
        self.user_combo.grid(row=0, column=1, sticky="ew", padx=(10, 0))
        self.user_combo.bind("<<ComboboxSelected>>",
                             lambda _e: self._on_user_change())
        user_actions = tk.Frame(user_frame, bg=self.card_bg)
        user_actions.grid(row=0, column=2, padx=(10, 0), sticky="e")
        ttk.Button(user_actions, text="Edit Details", command=self._edit_selected_user_details,
                   style="Secondary.TButton").pack(side="left")
        ttk.Button(user_actions, text="Security", command=self._change_selected_user_password,
                   style="Secondary.TButton").pack(side="left", padx=(6, 0))

        details = tk.Frame(user_frame, bg=self.card_bg)
        details.grid(row=1, column=0, columnspan=3, sticky="ew", pady=(15, 0))
        details.grid_columnconfigure(1, weight=1)

        profile_fields = [
            ("Full Name:", self.profile_name_var),
            ("Job Role:", self.profile_job_role_var),
            ("Required Hours:", self.profile_required_hours_var),
            ("Student ID:", self.profile_student_id_var),
            ("Course:", self.profile_course_var),
            ("University:", self.profile_university_var),
        ]

        for r, (label, var) in enumerate(profile_fields):
            tk.Label(details, text=label, bg=self.card_bg, fg=self.text_light, font=("Segoe UI", 9, "bold")).grid(row=r,
                                                                                                                  column=0,
                                                                                                                  sticky="w",
                                                                                                                  pady=3)
            tk.Label(details, textvariable=var, bg=self.card_bg, fg=self.text_dark, font=("Segoe UI", 10)).grid(row=r,
                                                                                                                column=1,
                                                                                                                sticky="w",
                                                                                                                padx=(
                                                                                                                    15, 0),
                                                                                                                pady=3)

        hours_frame = create_card(left_col, "OJT Progress")
        hours_frame.columnconfigure(0, weight=0)
        hours_frame.columnconfigure(1, weight=1)
        hours_frame.columnconfigure(2, weight=0)
        tk.Label(hours_frame, text="Set Goal (Hrs):", bg=self.card_bg, fg=self.text_light, font=("Segoe UI", 9)).grid(
            row=0, column=0, sticky="w")
        ttk.Entry(hours_frame, textvariable=self.required_hours_var, width=12, font=("Segoe UI", 10)).grid(row=0,
                                                                                                           column=1,
                                                                                                           padx=(
                                                                                                               10, 10),
                                                                                                           sticky="ew")
        ttk.Button(hours_frame, text="Apply", command=self._recompute_totals, style="Secondary.TButton").grid(row=0,
                                                                                                              column=2,
                                                                                                              sticky="w")

        cards_row = tk.Frame(hours_frame, bg=self.card_bg)
        cards_row.grid(row=1, column=0, columnspan=3,
                       sticky="ew", pady=(15, 12))
        for idx in range(3):
            cards_row.grid_columnconfigure(idx, weight=1)

        def mini_card(parent, col, value_var, caption):
            frame = tk.Frame(parent, bg="#F8F9FA", relief="flat", highlightthickness=1, highlightbackground=BORDER,
                             padx=12, pady=10)
            frame.grid(row=0, column=col, padx=(
                0 if col == 0 else 10, 0), sticky="ew")
            tk.Label(frame, textvariable=value_var, font=FONT_NUMBER,
                     bg="#F8F9FA", fg=PRIMARY).pack(anchor="center")
            tk.Label(frame, text=caption.upper(), font=FONT_SMALL, bg="#F8F9FA", fg=TEXT_LIGHT).pack(anchor="center",
                                                                                                     pady=(2, 0))

        mini_card(cards_row, 0, self.required_hours_display_var, "Required")
        mini_card(cards_row, 1, self.completed_hours_var, "Completed")
        mini_card(cards_row, 2, self.remaining_hours_var, "Remaining")

        progress_wrap = tk.Frame(hours_frame, bg=self.card_bg)
        progress_wrap.grid(row=3, column=0, columnspan=3,
                           sticky="ew", pady=(10, 0))
        progress_wrap.grid_columnconfigure(0, weight=1)
        self.progress_bar = ttk.Progressbar(
            progress_wrap,
            orient="horizontal",
            mode="determinate",
            maximum=100,
            variable=self.progress_value_var,
            style="Modern.Horizontal.TProgressbar",
        )
        self.progress_bar.grid(row=0, column=0, sticky="ew", ipady=2)

        self.progress_inside_label = tk.Label(progress_wrap, textvariable=self.progress_percent_var, bg=self.card_bg,
                                              fg=self.text_dark, font=("Segoe UI", 9, "bold"))
        self.progress_inside_label.place(relx=0.5, rely=0.5, anchor="center")

        milestone_row = tk.Frame(hours_frame, bg=self.card_bg)
        milestone_row.grid(row=4, column=0, columnspan=3,
                           sticky="ew", pady=(8, 0))
        for idx, txt in enumerate(("25%", "50%", "75%", "100%")):
            tk.Label(milestone_row, text=txt, bg=self.card_bg, fg=self.text_light, font=("Segoe UI", 8)).grid(row=0,
                                                                                                              column=idx,
                                                                                                              sticky="ew")
            milestone_row.grid_columnconfigure(idx, weight=1)

        self.status_label = tk.Label(hours_frame, textvariable=self.status_var, bg=self.card_bg, fg=self.primary_blue,
                                     font=("Segoe UI", 10, "bold"))
        self.status_label.grid(
            row=5, column=0, columnspan=3, sticky="w", pady=(12, 0))

        intern_summary_frame = create_card(left_col, "Active Interns")
        summary_header = tk.Frame(intern_summary_frame, bg=self.card_bg)
        summary_header.pack(fill="x", pady=(0, 10))
        tk.Label(summary_header, textvariable=self.intern_count_var, bg=self.card_bg, fg=self.text_dark,
                 font=("Segoe UI", 10, "bold")).pack(side="left")

        summary_cols = ("full_name", "job_role", "university", "status")
        self.summary_tree = ttk.Treeview(
            intern_summary_frame, columns=summary_cols, show="headings", height=8)
        self.summary_tree.heading("full_name", text="NAME")
        self.summary_tree.heading("job_role", text="ROLE")
        self.summary_tree.heading("university", text="UNIVERSITY")
        self.summary_tree.heading("status", text="STATUS")
        self.summary_tree.column("full_name", width=150, anchor="w")
        self.summary_tree.column("job_role", width=100, anchor="w")
        self.summary_tree.column("university", width=120, anchor="w")
        self.summary_tree.column("status", width=80, anchor="center")
        self.summary_tree.pack(fill="both", expand=True)

        self.summary_tree.bind("<Button-3>", self._show_summary_context_menu)
        self.summary_context_menu = tk.Menu(self.root, tearoff=0)
        self.summary_context_menu.add_command(label="Soft Delete",
                                              command=lambda: self._soft_delete_selected_intern("delete"))
        self.summary_context_menu.add_command(label="Restore",
                                              command=lambda: self._soft_delete_selected_intern("restore"))

        tk.Label(intern_summary_frame, textvariable=self.scan_feedback_var, bg=self.card_bg, fg=self.accent_green,
                 font=("Segoe UI", 9, "italic")).pack(anchor="w", pady=(10, 0))

        record_frame = create_card(
            right_col, "Attendance History", expand=True)
        control_row = tk.Frame(record_frame, bg=self.card_bg)
        control_row.pack(fill="x", pady=(0, 12))

        row2 = tk.Frame(control_row, bg=self.card_bg)
        row2.pack(fill="x")
        tk.Label(row2, text="Filters:", bg=self.card_bg, fg=self.text_light, font=("Segoe UI", 9, "bold")).pack(
            side="left")

        ttk.Button(row2, text="Start Date", command=self._select_from_date, style="Secondary.TButton").pack(side="left",
                                                                                                            padx=(8, 0))
        tk.Entry(row2, textvariable=self.from_date_var, width=12, state="readonly", font=("Segoe UI", 9), bd=1,
                 relief="solid").pack(side="left", padx=(4, 0))

        ttk.Button(row2, text="End Date", command=self._select_to_date, style="Secondary.TButton").pack(side="left",
                                                                                                        padx=(10, 0))
        tk.Entry(row2, textvariable=self.to_date_var, width=12, state="readonly", font=("Segoe UI", 9), bd=1,
                 relief="solid").pack(side="left", padx=(4, 0))

        filter_actions = tk.Frame(row2, bg=self.card_bg)
        filter_actions.pack(side="left", padx=(15, 0))
        ttk.Button(filter_actions, text="Filter", command=self._apply_record_filters, style="Action.TButton").pack(
            side="left")
        ttk.Button(filter_actions, text="Reset", command=self._reset_record_filters, style="Secondary.TButton").pack(
            side="left", padx=(6, 0))

        tk.Label(row2, textvariable=self.summary_var, bg="#F8F9FA", fg=self.primary_blue, font=("Segoe UI", 9, "bold"),
                 padx=8, pady=4).pack(side="right")

        table_view = tk.Frame(record_frame, bg=self.card_bg)
        table_view.pack(fill="both", expand=True)
        table_view.grid_rowconfigure(0, weight=1)
        table_view.grid_columnconfigure(0, weight=1)

        columns = ("date", "time_in", "time_out", "hours", "status")
        self.tree = ttk.Treeview(
            table_view, columns=columns, show="headings", height=18)
        self.tree.heading("date", text="DATE")
        self.tree.heading("time_in", text="TIME IN")
        self.tree.heading("time_out", text="TIME OUT")
        self.tree.heading("hours", text="HOURS")
        self.tree.heading("status", text="STATUS")

        self.tree.column("date", width=120, anchor="center")
        self.tree.column("time_in", width=120, anchor="center")
        self.tree.column("time_out", width=120, anchor="center")
        self.tree.column("hours", width=100, anchor="center")
        self.tree.column("status", width=120, anchor="center")

        # Alternating rows
        self.tree.tag_configure("odd", background="#F9FBFD")
        self.tree.tag_configure("even", background="#FFFFFF")
        # Status colors
        self.tree.tag_configure("late", foreground=DANGER)
        self.tree.tag_configure("undertime", foreground=WARNING)
        self.tree.tag_configure("half_day", foreground=WARNING)
        self.tree.tag_configure("on_time", foreground=SUCCESS)
        self.tree.tag_configure("open", foreground=TEXT_LIGHT)

        self.tree.grid(row=0, column=0, sticky="nsew")
        self.tree.bind("<Configure>", self._resize_tree_columns)
        self.tree.bind("<Button-3>", self._show_context_menu)

        scrollbar = ttk.Scrollbar(
            table_view, orient="vertical", command=self.tree.yview)
        scrollbar.grid(row=0, column=1, sticky="ns")
        self.tree.configure(yscrollcommand=scrollbar.set)
        self.empty_state_label = tk.Label(
            table_view,
            text="No attendance records yet.",
            bg="white",
            fg="#60708a",
            font=("Segoe UI", 10, "italic"),
        )
        self.empty_state_label.place(relx=0.5, rely=0.5, anchor="center")

        self.context_menu = tk.Menu(self.root, tearoff=0)
        self.add_submenu = tk.Menu(self.context_menu, tearoff=0)
        self.add_submenu.add_command(
            label="Add Full Entry", command=self._add_entry_from_context)
        self.add_submenu.add_command(
            label="Add Date", command=self._add_date_only_from_context)
        self.add_submenu.add_command(
            label="Add Time In", command=self._add_time_in_only_from_context)
        self.add_submenu.add_command(
            label="Add Time Out", command=self._add_time_out_only_from_context)
        self.context_menu.add_cascade(label="Add Entry", menu=self.add_submenu)
        self.context_menu.add_separator()
        self.edit_submenu = tk.Menu(self.context_menu, tearoff=0)
        self.edit_submenu.add_command(
            label="Edit Date", command=self._edit_selected_date)
        self.edit_submenu.add_command(
            label="Edit Time In", command=self._edit_selected_time_in)
        self.edit_submenu.add_command(
            label="Edit Time Out", command=self._edit_selected_time_out)
        self.context_menu.add_cascade(
            label="Edit Entry", menu=self.edit_submenu)
        self.context_menu.add_command(
            label="Delete Entry", command=self._delete_selected)

        bottom_actions = tk.Frame(container, bg=BG)
        bottom_actions.pack(side="bottom", fill="x", pady=(10, 5))

        scan_btn = ttk.Button(
            bottom_actions,
            text="📷 Scan QR",
            command=self._scan_qr_attendance,
            style="Secondary.TButton"
        )
        scan_btn.pack(anchor="center", pady=(8, 12))

# Optional: store reference
        self.dashboard_scan_btn = scan_btn

        footer = tk.Label(
            container,
            text="© 2026 AUTOLIV Internship Management System    Version 1.0",
            font=("Segoe UI", 9),
            bg="#f3f6fb",
            fg="#7b8799",
        )
        footer.pack(anchor="e", pady=(6, 0))

    def _on_btn_enter(self, event, hover_style=None, widget=None):
        try:
            w = widget or event.widget
            if hover_style and isinstance(w, ttk.Button):
                w.configure(style=hover_style)
            else:
                w.configure(bg=self.primary_blue)
        except Exception:
            pass

    def _on_btn_leave(self, event, normal_style=None, widget=None):
        try:
            w = widget or event.widget
            if normal_style and isinstance(w, ttk.Button):
                w.configure(style=normal_style)
        except Exception:
            pass

    def _make_button(self, parent, text, command=None, style="Secondary.TButton", pack_opts=None):
        btn = ttk.Button(parent, text=text, command=command, style=style)
        hover_style = None
        if style == "Action.TButton":
            hover_style = "ActionHover.TButton"
        elif style == "Secondary.TButton":
            hover_style = "SecondaryHover.TButton"

        if hover_style:
            btn.bind("<Enter>", lambda e, s=hover_style,
                     w=btn: self._on_btn_enter(e, hover_style=s, widget=w))
            btn.bind("<Leave>", lambda e, n=style,
                     w=btn: self._on_btn_leave(e, normal_style=n, widget=w))

        if pack_opts:
            btn.pack(**pack_opts)
        return btn

    def _ensure_default_user(self):
        self.users_data = {k: self._normalize_user_payload(
            k, v) for k, v in self.users_data.items()}
        if "Default" in self.users_data and self._is_auto_default_user("Default", self.users_data.get("Default")):
            self.users_data.pop("Default", None)
        if self.users_data:
            active_users = self._active_user_keys()
            if self.active_user_var.get() not in active_users:
                self.active_user_var.set(
                    active_users[0] if active_users else "")
            self._last_active_user = self.active_user_var.get()
            self._refresh_user_combo()
            return

        self.active_user_var.set("")
        self._last_active_user = ""
        self._refresh_user_combo()
        self._save_records()

    def _refresh_user_combo(self):
        users = sorted(self._active_user_keys(),
                       key=lambda key: self._full_name_for_user(key).lower())
        self._user_to_display = {
            user: self._display_label_for_user(user) for user in users}
        self._display_to_user = {display: user for user,
                                 display in self._user_to_display.items()}
        self.user_combo["values"] = [self._user_to_display[user]
                                     for user in users]
        if users and self.active_user_var.get() not in users:
            self.active_user_var.set(users[0])
        active_key = self.active_user_var.get().strip()
        self.active_user_display_var.set(
            self._user_to_display.get(active_key, ""))
        self._refresh_profile_details()
        self._refresh_intern_summary()

    def _active_user(self):
        return self.active_user_var.get().strip()

    def _commit_user_from_view(self, user: str):
        if not user or user not in self.users_data:
            return
        self.users_data[user]["required_hours"] = self.required_hours_var.get(
        ).strip()
        # Only commit from tree when we're not showing a filtered view (tree = full data).
        if not self._filters_active:
            self.users_data[user]["records"] = self._tree_records()

    def _refresh_profile_details(self):
        user = self._active_user()
        data = self.users_data.get(user, {})
        self.profile_name_var.set(
            self._full_name_for_user(user) if user else "-")
        self.profile_job_role_var.set(data.get("job_role", "") or "-")
        self.profile_required_hours_var.set(
            str(data.get("required_hours", "500")))
        self.profile_student_id_var.set(data.get("student_id", "") or "-")
        self.profile_course_var.set(data.get("course", "") or "-")
        self.profile_university_var.set(data.get("university", "") or "-")

    def _compute_user_completed_hours(self, user):
        completed = 0.0
        for row in self.users_data.get(user, {}).get("records", []):
            if len(row) < 4 or not row[3]:
                continue
            try:
                completed += float(row[3])
            except Exception:
                continue
        return completed

    def _refresh_intern_summary(self):
        if not hasattr(self, "summary_tree"):
            return
        for item in self.summary_tree.get_children():
            self.summary_tree.delete(item)
        active_count = 0
        for user in sorted(self.users_data.keys(), key=lambda key: self._full_name_for_user(key).lower()):
            if user == "Default":
                continue
            data = self.users_data.get(user, {})
            is_deleted = bool(data.get("is_deleted"))
            if not is_deleted:
                active_count += 1
            status = "Deleted" if is_deleted else "Active"
            self.summary_tree.insert(
                "",
                "end",
                iid=user,
                values=(
                    self._full_name_for_user(user),
                    data.get("job_role", ""),
                    data.get("university", ""),
                    status,
                ),
            )
        self.intern_count_var.set(f"Active Interns: {active_count}")

    def _show_summary_context_menu(self, event):
        item = self.summary_tree.identify_row(event.y)
        if not item:
            return
        self.summary_tree.selection_set(item)
        self.summary_tree.focus(item)
        if not self.is_admin:
            self.summary_context_menu.entryconfig(0, state="disabled")
            self.summary_context_menu.entryconfig(1, state="disabled")
            self.summary_context_menu.tk_popup(event.x_root, event.y_root)
            return
        data = self.users_data.get(item, {})
        is_deleted = bool(data.get("is_deleted"))
        self.summary_context_menu.entryconfig(
            0, state=("disabled" if is_deleted else "normal"))
        self.summary_context_menu.entryconfig(
            1, state=("normal" if is_deleted else "disabled"))
        self.summary_context_menu.tk_popup(event.x_root, event.y_root)

    def _open_user_profile_window(self):
        # Always use Toplevel (never create a second Tk()).
        if self.user_profile_window is not None and getattr(self.user_profile_window, 'winfo_exists', lambda: False)():
            try:
                self.user_profile_window.deiconify()
                self.user_profile_window.lift()
            except Exception:
                pass
            self._refresh_user_profile_window()
            return

        win = tk.Toplevel(self.root)
        win.title("USER PROFILE - OJT Tracker")
        try:
            win.geometry("1000x750")
            win.state("zoomed")
        except Exception:
            pass
        win.configure(bg=self.bg_color)

        self.user_profile_window = win
        self.user_profile_widgets = {}

        container = tk.Frame(win, bg=self.bg_color, padx=25, pady=25)
        container.pack(fill="both", expand=True)

        header = tk.Frame(container, bg=self.bg_color)
        header.pack(fill="x", pady=(0, 20))
        tk.Label(header, text="INTERN DASHBOARD", font=("Segoe UI", 24, "bold"), bg=self.bg_color,
                 fg=self.primary_blue).pack(side="left")
        header_actions = tk.Frame(header, bg=self.bg_color)
        header_actions.pack(side="right")
        self._make_button(header_actions, "Back to Dashboard", command=self._back_to_dashboard_from_profile,
                          style="Secondary.TButton", pack_opts={"side": "right"})

        content = tk.Frame(container, bg=self.bg_color)
        content.pack(fill="both", expand=True)
        content.grid_rowconfigure(0, weight=1)
        content.grid_columnconfigure(0, weight=3, uniform="profile")
        content.grid_columnconfigure(1, weight=4, uniform="profile")

        left_col = tk.Frame(content, bg=self.bg_color)
        left_col.grid(row=0, column=0, sticky="nsew", padx=(0, 12))
        right_col = tk.Frame(content, bg=self.bg_color)
        right_col.grid(row=0, column=1, sticky="nsew", padx=(12, 0))

        def create_card(parent, heading, expand=False):
            card_outer = tk.Frame(parent, bg="#E9EEF6", padx=10, pady=8)
            card_outer.pack(fill="both" if expand else "x",
                            expand=expand, pady=8)
            card = tk.Frame(card_outer, bg=self.card_bg,
                            highlightthickness=1, highlightbackground=self.border_color)
            card.pack(fill="both", expand=True, padx=4, pady=4)

            h_frame = tk.Frame(card, bg="#F8F9FA")
            h_frame.pack(fill="x")
            tk.Label(h_frame, text=heading, font=("Segoe UI", 11, "bold"), bg="#F8F9FA", fg=self.text_dark).pack(
                anchor="w", padx=15, pady=10)
            tk.Frame(card, height=1, bg=self.border_color).pack(fill="x")

            body = tk.Frame(card, bg=self.card_bg, padx=20, pady=20)
            body.pack(fill="both", expand=True)
            return body

        intern_card = create_card(left_col, "INTERN INFORMATION")
        intern_card.columnconfigure(1, weight=1)
        fields = [
            ("Full Name", "full_name"),
            ("Job Role", "job_role"),
            ("Required Hours", "required_hours"),
            ("Student ID", "student_id"),
            ("Course", "course"),
            ("University", "university"),
        ]
        for idx, (label, key) in enumerate(fields):
            tk.Label(intern_card, text=label + ":", bg=self.card_bg, fg=self.text_light,
                     font=("Segoe UI", 10, "bold")).grid(row=idx, column=0, sticky="w", pady=5)
            value_lbl = tk.Label(
                intern_card, text="-", bg=self.card_bg, fg=self.text_dark, font=("Segoe UI", 10))
            value_lbl.grid(row=idx, column=1, sticky="w", padx=(15, 0), pady=5)
            self.user_profile_widgets[key] = value_lbl

        progress_card = create_card(left_col, "PROGRESS OVERVIEW")
        metrics = tk.Frame(progress_card, bg=self.card_bg)
        metrics.pack(fill="x")
        for idx in range(2):
            metrics.grid_columnconfigure(idx, weight=1)
            metrics.grid_rowconfigure(idx, weight=1)

        def metric(r, c, caption):
            panel = tk.Frame(metrics, bg="#F8F9FA", highlightthickness=1, highlightbackground=self.border_color,
                             padx=10, pady=12)
            panel.grid(row=r, column=c, sticky="ew", padx=5, pady=5)
            value_lbl = tk.Label(panel, text="0.00", font=(
                "Segoe UI", 16, "bold"), bg="#F8F9FA", fg=self.primary_blue)
            value_lbl.pack()
            tk.Label(panel, text=caption.upper(), font=("Segoe UI", 8, "bold"), bg="#F8F9FA", fg=self.text_light).pack(
                pady=(2, 0))
            return value_lbl

        self.user_profile_widgets["metric_required"] = metric(0, 0, "Required")
        self.user_profile_widgets["metric_completed"] = metric(
            0, 1, "Completed")
        self.user_profile_widgets["metric_remaining"] = metric(
            1, 0, "Remaining")
        self.user_profile_widgets["metric_progress"] = metric(
            1, 1, "Progress %")

        self.user_profile_widgets["status"] = tk.Label(progress_card, text="Status: -", bg=self.card_bg,
                                                       fg=self.primary_blue, font=("Segoe UI", 10, "bold"))
        self.user_profile_widgets["status"].pack(anchor="w", pady=(15, 0))

        progress_wrap = tk.Frame(progress_card, bg=self.card_bg)
        progress_wrap.pack(fill="x", pady=(10, 0))
        progress_canvas = tk.Canvas(
            progress_wrap, height=20, bg="#E8EAED", highlightthickness=0)
        progress_canvas.pack(fill="x")
        self.user_profile_widgets["progress_canvas"] = progress_canvas
        self.user_profile_widgets["progress_text"] = tk.Label(progress_wrap, text="0.00%", bg=self.card_bg,
                                                              fg=self.text_dark, font=("Segoe UI", 9, "bold"))
        self.user_profile_widgets["progress_text"].pack(
            anchor="e", pady=(4, 0))

        attendance_card = create_card(right_col, "ATTENDANCE LOG", expand=True)
        cols = ("date", "time_in", "time_out", "hours", "status")
        profile_tree = ttk.Treeview(
            attendance_card, columns=cols, show="headings", height=20)
        profile_tree.heading("date", text="DATE")
        profile_tree.heading("time_in", text="TIME IN")
        profile_tree.heading("time_out", text="TIME OUT")
        profile_tree.heading("hours", text="HOURS")
        profile_tree.heading("status", text="STATUS")

        profile_tree.column("date", width=120, anchor="center")
        profile_tree.column("time_in", width=120, anchor="center")
        profile_tree.column("time_out", width=120, anchor="center")
        profile_tree.column("hours", width=100, anchor="center")
        profile_tree.column("status", width=120, anchor="center")

        profile_tree.tag_configure("late", foreground=DANGER)
        profile_tree.tag_configure("undertime", foreground=WARNING)
        profile_tree.tag_configure("half_day", foreground=WARNING)
        profile_tree.tag_configure("on_time", foreground=SUCCESS)
        profile_tree.tag_configure("open", foreground=TEXT_LIGHT)

        profile_tree.pack(side="left", fill="both", expand=True)
        scrollbar = ttk.Scrollbar(
            attendance_card, orient="vertical", command=profile_tree.yview)
        scrollbar.pack(side="right", fill="y")
        profile_tree.configure(yscrollcommand=scrollbar.set)
        self.user_profile_widgets["tree"] = profile_tree

        def on_close():
            try:
                win.destroy()
            except Exception:
                pass
            self.user_profile_window = None
            self.user_profile_widgets = {}

        try:
            win.protocol("WM_DELETE_WINDOW", on_close)
        except Exception:
            pass
        self._refresh_user_profile_window()

    def _back_to_dashboard_from_profile(self):
        try:
            if self.main_ui_window is None or not getattr(self.main_ui_window, 'winfo_exists', lambda: False)():
                try:
                    self._open_main_ui_window()
                except Exception:
                    pass
            if self.main_ui_window is not None and getattr(self.main_ui_window, 'winfo_exists', lambda: False)():
                try:
                    self.main_ui_window.deiconify()
                    try:
                        self.main_ui_window.state("zoomed")
                    except Exception:
                        pass
                    self.main_ui_window.lift()
                    self.main_ui_window.focus_force()
                except Exception:
                    pass
        except Exception:
            pass

        try:
            if self.user_profile_window is not None and getattr(self.user_profile_window, 'winfo_exists', lambda: False)():
                try:
                    self.user_profile_window.destroy()
                except Exception:
                    try:
                        self.user_profile_window.withdraw()
                    except Exception:
                        pass
        except Exception:
            pass
        self.user_profile_window = None
        self.user_profile_widgets = {}

    def _open_main_ui_window(self):
        if self.main_ui_window is not None and self.main_ui_window.winfo_exists():
            self.main_ui_window.deiconify()
            self.main_ui_window.state("zoomed")
            self.main_ui_window.lift()
            self._refresh_main_ui_window()
            return

        win = tk.Toplevel(self.root)
        win.title("AUTOLIV - Internship Management System")
        win.geometry("1100x750")
        win.state("zoomed")
        win.configure(bg=self.bg_color)
        self.main_ui_window = win
        self.main_ui_widgets = {}

        container = tk.Frame(win, bg=self.bg_color, padx=30, pady=30)
        container.pack(fill="both", expand=True)

        right_frame = tk.Frame(container, bg=self.bg_color)
        right_frame.pack(fill="both", expand=True)

        top = tk.Frame(right_frame, bg=self.bg_color)
        top.pack(fill="x", pady=(0, 20))

        brand_frame = tk.Frame(top, bg=self.bg_color)
        brand_frame.pack(side="left")
        tk.Label(brand_frame, text="🔷", font=("Segoe UI", 22), bg=self.bg_color,
                 fg=self.primary_blue).pack(side="left", padx=(0, 8))
        tk.Label(brand_frame, text="AUTOLIV", font=("Segoe UI", 28, "bold"), bg=self.bg_color,
                 fg=self.primary_blue).pack(side="left")
        tk.Label(brand_frame, text="Internship Management System", font=("Segoe UI", 16), bg=self.bg_color,
                 fg=self.text_light).pack(side="left", padx=(10, 0), pady=(8, 0))

        right_actions = tk.Frame(top, bg=self.bg_color)
        right_actions.pack(side="right", anchor="ne")
        self._make_button(right_actions, "User Login", command=self._user_login_from_main,
                          style="Action.TButton", pack_opts={"side": "left"})
        admin_btn = self._make_button(
            right_actions,
            self._main_admin_button_text(),
            command=self._open_admin_ui_from_main,
            style="Secondary.TButton",
            pack_opts={"side": "left", "padx": (10, 0)}
        )
        self.main_ui_widgets["admin_btn"] = admin_btn

        table_card = tk.Frame(right_frame, bg=self.card_bg, bd=0, highlightthickness=1,
                              highlightbackground=self.border_color)
        table_card.pack(fill="both", expand=True)

        # Single scan section (no duplicates)
        bottom_section = tk.Frame(right_frame, bg=self.bg_color)
        bottom_section.pack(fill="x", pady=(15, 20))
        scan_btn = ttk.Button(
            bottom_section,
            text="📷 Scan QR",
            command=self._scan_qr_attendance,
            style="Secondary.TButton"
        )
        scan_btn.pack(anchor="center")
        self.main_qr_btn = scan_btn

        header_canvas = tk.Canvas(table_card, height=68, highlightthickness=0)
        header_canvas.pack(fill="x")

        def _draw_header_gradient(event=None):
            w = header_canvas.winfo_width() or 800
            h = header_canvas.winfo_height() or 68
            header_canvas.delete("grad")
            c1 = self._hex_to_rgb(self.primary_blue)
            c2 = self._hex_to_rgb(self.secondary_blue)
            for i in range(w):
                r = int(c1[0] + (c2[0]-c1[0]) * (i / max(w-1, 1)))
                g = int(c1[1] + (c2[1]-c1[1]) * (i / max(w-1, 1)))
                b = int(c1[2] + (c2[2]-c1[2]) * (i / max(w-1, 1)))
                header_canvas.create_line(
                    i, 0, i, h, fill=self._rgb_to_hex((r, g, b)), tags=("grad",))
            header_canvas.create_rectangle(
                0, 0, w, h, fill="", outline="", tags=("grad",))
        header_canvas.bind("<Configure>", _draw_header_gradient)
        header_canvas.create_text(18, 34, anchor="w", text="ATTENDANCE RECORD OVERVIEW", font=(
            "Segoe UI", 12, "bold"), fill="#FFFFFF", tags=("label",))
        tk.Frame(table_card, height=1, bg=self.border_color).pack(fill="x")

        table_wrap = tk.Frame(table_card, bg=self.card_bg, padx=20, pady=20)
        table_wrap.pack(fill="both", expand=True)

        summary_row = tk.Frame(table_wrap, bg=self.card_bg)
        summary_row.pack(fill="x", pady=(0, 12))
        self.main_summary_vars = {
            "total": tk.StringVar(value="0"),
            "present": tk.StringVar(value="0"),
            "late": tk.StringVar(value="0"),
            "absent": tk.StringVar(value="0"),
        }

        def small_card(parent, var, caption):
            f = tk.Frame(parent, bg="#F8F9FA", padx=12, pady=10,
                         highlightthickness=1, highlightbackground=self.border_color)
            tk.Label(f, textvariable=var, font=("Segoe UI", 18, "bold"),
                     bg="#F8F9FA", fg=self.primary_blue).pack()
            tk.Label(f, text=caption, font=("Segoe UI", 9, "bold"),
                     bg="#F8F9FA", fg=self.text_light).pack()
            return f
        small_card(summary_row, self.main_summary_vars["total"], "Total Interns").pack(
            side="left", expand=True, fill="x", padx=(0, 10))
        small_card(summary_row, self.main_summary_vars["present"], "Present Today").pack(
            side="left", expand=True, fill="x", padx=(0, 10))
        small_card(summary_row, self.main_summary_vars["late"], "Late").pack(
            side="left", expand=True, fill="x", padx=(0, 10))
        small_card(summary_row, self.main_summary_vars["absent"], "Absent").pack(
            side="left", expand=True, fill="x")

        search_frame = tk.Frame(table_wrap, bg=self.card_bg)
        search_frame.pack(fill="x", pady=(0, 10))
        self.main_search_var = tk.StringVar(value="")
        tk.Entry(search_frame, textvariable=self.main_search_var, width=28, font=("Segoe UI", 10)).pack(
            side="right", padx=(6, 0))
        tk.Label(search_frame, text="Search:", bg=self.card_bg,
                 fg=self.text_light, font=("Segoe UI", 10)).pack(side="right")

        def _on_search_key(_e=None):
            self._refresh_main_ui_window()
        win.bind("<KeyRelease>", _on_search_key)

        cols = ("name", "time_in", "time_out", "date")
        tree = ttk.Treeview(table_wrap, columns=cols,
                            show="headings", height=20)
        tree.heading("name", text="INTERN NAME")
        tree.heading("time_in", text="TIME IN")
        tree.heading("time_out", text="TIME OUT")
        tree.heading("date", text="DATE")

        tree.column("name", width=300, anchor="w")
        tree.column("time_in", width=150, anchor="center")
        tree.column("time_out", width=150, anchor="center")
        tree.column("date", width=150, anchor="center")

        tree.tag_configure("late", foreground=DANGER)
        tree.tag_configure("half_day", foreground=WARNING)
        tree.tag_configure("on_time", foreground=SUCCESS)

        tree.pack(side="left", fill="both", expand=True)
        scrollbar = ttk.Scrollbar(
            table_wrap, orient="vertical", command=tree.yview)
        scrollbar.pack(side="right", fill="y")
        tree.configure(yscrollcommand=scrollbar.set)
        self.main_ui_widgets["tree"] = tree

        def on_close():
            self.main_ui_window = None
            self.main_ui_widgets = {}
            win.destroy()

        win.protocol("WM_DELETE_WINDOW", on_close)
        self._refresh_main_ui_window()

    def _refresh_main_ui_window(self):
        if self.main_ui_window is None or not self.main_ui_window.winfo_exists():
            return
        tree = self.main_ui_widgets.get("tree")
        admin_btn = self.main_ui_widgets.get("admin_btn")
        if admin_btn is not None:
            admin_btn.configure(text=self._main_admin_button_text())
        if tree is None:
            return
        for item in tree.get_children():
            tree.delete(item)
        rows = []
        total = 0
        present_set = set()
        late_set = set()
        today_str = date.today().strftime(DATE_FMT)
        q = (getattr(self, "main_search_var",
             tk.StringVar()).get() or "").strip().lower()

        for user, data in self.users_data.items():
            if data.get("is_deleted"):
                continue
            total += 1
            display_name = self._full_name_for_user(user)
            matches_query = True
            if q:
                matches_query = q in display_name.lower()

            for rec in data.get("records", []):
                if len(rec) < 4:
                    continue
                rec_date = rec[0]
                rec_time_in = rec[1]
                rec_time_out = rec[2]
                status_val = (rec[4] if len(rec) > 4 else "")

                if matches_query:
                    rows.append((display_name, rec_time_in, rec_time_out,
                                 rec_date, status_val))

                try:
                    if str(rec_date).strip() == today_str and rec_time_in:
                        present_set.add(user)
                        if status_val and str(status_val).strip().lower().startswith("late"):
                            late_set.add(user)
                except Exception:
                    pass

        def sort_key(r):
            try:
                d = datetime.strptime(r[3], DATE_FMT)
            except Exception:
                d = datetime.min
            return d

        rows.sort(key=sort_key, reverse=True)
        for row in rows:
            values = row[:4]
            status_val = row[4] if len(row) > 4 else ""
            tag = "on_time"
            try:
                if status_val and str(status_val).strip().lower().startswith("late"):
                    tag = "late"
                elif status_val and str(status_val).strip().lower().startswith("half"):
                    tag = "half_day"
                else:
                    tag = "on_time"
            except Exception:
                tag = "on_time"
            tree.insert("", "end", values=values, tags=(tag,))

        if hasattr(self, "main_summary_vars"):
            try:
                present = len(present_set)
                late = len(late_set)
                self.main_summary_vars["total"].set(str(total))
                self.main_summary_vars["present"].set(str(present))
                self.main_summary_vars["late"].set(str(late))
                self.main_summary_vars["absent"].set(
                    str(max(total - present, 0)))
            except Exception:
                pass

    def _user_login_from_main(self):
        parent = self.main_ui_window if (
            self.main_ui_window and self.main_ui_window.winfo_exists()) else self.root
        active_users = sorted(self._active_user_keys(),
                              key=lambda key: self._full_name_for_user(key).lower())
        if not active_users:
            messagebox.showerror(
                "User Login", "No active users found.", parent=parent)
            return

        dialog = tk.Toplevel(parent)
        dialog.title("User Login")
        dialog.transient(parent)
        dialog.grab_set()
        dialog.resizable(False, False)
        form = ttk.Frame(dialog, padding=12)
        form.pack(fill="both", expand=True)

        user_display_map = {user: self._display_label_for_user(
            user) for user in active_users}
        display_to_user = {display: user for user,
                           display in user_display_map.items()}
        selected_display_var = tk.StringVar(
            value=user_display_map[active_users[0]])
        password_var = tk.StringVar()
        result = {"user": None}

        ttk.Label(form, text="User:").grid(
            row=0, column=0, sticky="w", pady=(0, 8))
        user_combo = ttk.Combobox(
            form,
            textvariable=selected_display_var,
            values=[user_display_map[user] for user in active_users],
            state="readonly",
            width=34,
        )
        user_combo.grid(row=0, column=1, sticky="ew", padx=(8, 0), pady=(0, 8))

        ttk.Label(form, text="Password:").grid(row=1, column=0, sticky="w")
        password_entry = ttk.Entry(
            form, textvariable=password_var, show="*", width=34)
        password_entry.grid(row=1, column=1, sticky="ew", padx=(8, 0))
        form.grid_columnconfigure(1, weight=1)

        def on_login():
            selected_display = selected_display_var.get().strip()
            user_key = display_to_user.get(selected_display, "")
            if not user_key:
                messagebox.showerror(
                    "User Login", "Select a valid user.", parent=dialog)
                return
            stored_password = str(self.users_data.get(
                user_key, {}).get("password", "")).strip()
            if not stored_password:
                messagebox.showerror("User Login", "This user has no password set. Ask admin to set one.",
                                     parent=dialog)
                return
            entered = password_var.get().strip()
            if not _verify_password(entered, stored_password):
                messagebox.showerror(
                    "User Login", "Incorrect password.", parent=dialog)
                return

            # If the stored password was legacy plaintext, upgrade it after a successful login.
            if stored_password and not _is_hashed_password(stored_password):
                self.users_data[user_key]["password"] = _hash_password(entered)
                self._save_records()

            result["user"] = user_key
            dialog.destroy()

        actions = ttk.Frame(form)
        actions.grid(row=2, column=0, columnspan=2,
                     sticky="e", pady=(10, 0))
        ttk.Button(actions, text="Cancel",
                   command=dialog.destroy).pack(side="right")
        ttk.Button(actions, text="Login", command=on_login).pack(
            side="right", padx=(0, 8))

        password_entry.focus_set()
        self._center_toplevel(dialog)
        self.root.wait_window(dialog)
        if not result["user"]:
            return
        self.active_user_var.set(result["user"])
        self._refresh_user_combo()
        self._last_active_user = result["user"]
        self._load_active_user_into_view()
        self._open_user_profile_window()
        try:
            if self.main_ui_window is not None and getattr(self.main_ui_window, 'winfo_exists', lambda: False)():
                try:
                    self.main_ui_window.withdraw()
                except Exception:
                    pass
            if self.user_profile_window is not None and getattr(self.user_profile_window, 'winfo_exists', lambda: False)():
                try:
                    self.user_profile_window.deiconify()
                    try:
                        self.user_profile_window.state("zoomed")
                    except Exception:
                        pass
                    self.user_profile_window.lift()
                    self.user_profile_window.focus_force()
                except Exception:
                    pass
            try:
                self.root.withdraw()
            except Exception:
                pass
        except Exception:
            pass

    def _open_admin_ui_from_main(self):
        if not self.is_admin:
            parent = self.main_ui_window if (
                self.main_ui_window and self.main_ui_window.winfo_exists()) else self.root
            password = simpledialog.askstring(
                "Admin Login", "Enter admin password:", show="*", parent=parent)
            if password is None:
                return
            if password != self.admin_password:
                messagebox.showerror(
                    "Access Denied", "Incorrect admin password.", parent=parent)
                return
            self.is_admin = True

        self.admin_btn.configure(text=self._effective_admin_button_text())
        if hasattr(self, "columns_frame"):
            self.admin_tools_frame.pack(
                fill="x", pady=(0, 4), before=self.columns_frame)
        else:
            self.admin_tools_frame.pack(fill="x", pady=(0, 4))
        self.admin_tools_frame.lift()
        self.admin_tools_frame.update_idletasks()
        self._refresh_intern_summary()
        self._refresh_main_ui_window()
        self.root.deiconify()
        self.root.state("zoomed")
        self.root.lift()
        self.root.focus_force()
        if self.main_ui_window is not None and self.main_ui_window.winfo_exists():
            self.main_ui_window.withdraw()

    def _refresh_user_profile_window(self):
        if self.user_profile_window is None or not self.user_profile_window.winfo_exists():
            return
        user = self._active_user()
        if not user or user not in self.users_data:
            return
        data = self.users_data[user]
        required = float(data.get("required_hours", "500") or 500)
        completed = self._compute_user_completed_hours(user)
        remaining = max(required - completed, 0.0)
        progress = 0.0 if required == 0 else min(
            (completed / required) * 100.0, 100.0)

        if progress < 30:
            status = "Just Started"
            color = self.accent_red
        elif progress < 70:
            status = "In Progress"
            color = self.accent_yellow
        elif progress < 100:
            status = "On Track"
            color = self.primary_blue
        else:
            status = "Completed"
            color = self.accent_green

        self.user_profile_widgets["full_name"].configure(
            text=self._full_name_for_user(user))
        self.user_profile_widgets["job_role"].configure(
            text=data.get("job_role", "") or "-")
        self.user_profile_widgets["required_hours"].configure(
            text=f"{required:.2f}")
        self.user_profile_widgets["student_id"].configure(
            text=data.get("student_id", "") or "-")
        self.user_profile_widgets["course"].configure(
            text=data.get("course", "") or "-")
        self.user_profile_widgets["university"].configure(
            text=data.get("university", "") or "-")

        self.user_profile_widgets["metric_required"].configure(
            text=f"{required:.2f}")
        self.user_profile_widgets["metric_completed"].configure(
            text=f"{completed:.2f}")
        self.user_profile_widgets["metric_remaining"].configure(
            text=f"{remaining:.2f}")
        self.user_profile_widgets["metric_progress"].configure(
            text=f"{progress:.2f}%")

        self.user_profile_widgets["status"].configure(
            text=f"Status: {status}", fg=color)

        canvas = self.user_profile_widgets.get("progress_canvas")
        if canvas is not None:
            try:
                self._animate_canvas_progress(canvas, progress, color)
                canvas.delete("ticks")
                width = canvas.winfo_width() or 400
                height = canvas.winfo_height() or 20
                for pct in (25, 50, 75):
                    tick_x = int(width * (pct / 100.0))
                    canvas.create_line(
                        tick_x, 0, tick_x, height, fill="white", width=1, dash=(2, 2), tags=("ticks",))
            except Exception:
                pass

        self.user_profile_widgets["progress_text"].configure(
            text=f"{progress:.2f}% Complete", fg=color)

        tree = self.user_profile_widgets["tree"]
        for item in tree.get_children():
            tree.delete(item)
        for row in data.get("records", []):
            if len(row) >= 5:
                tag = self._row_tag_for_status(row[4])
                tree.insert("", "end", values=row[:5], tags=(tag,))

    def _on_user_change(self):
        self._commit_user_from_view(self._last_active_user)
        selected_display = self.active_user_display_var.get().strip()
        selected_key = self._display_to_user.get(selected_display, "")
        if selected_key:
            self.active_user_var.set(selected_key)
        self._save_records()
        self._last_active_user = self._active_user()
        self._load_active_user_into_view()

    def _load_admin_password(self):
        if ADMIN_CONFIG_FILE.exists():
            try:
                with ADMIN_CONFIG_FILE.open("r", encoding="utf-8") as file:
                    payload = json.load(file)
                value = payload.get("admin_password", "").strip()
                if value:
                    return value
            except Exception:
                pass
        return DEFAULT_ADMIN_PASSWORD

    def _save_admin_password(self):
        payload = {"admin_password": self.admin_password}
        with ADMIN_CONFIG_FILE.open("w", encoding="utf-8") as file:
            json.dump(payload, file, indent=2)

    def _toggle_admin(self):
        if self.is_admin:
            messagebox.showinfo(
                "Admin", "Admin is already logged in. Use the top 'Admin Log Out' button to sign out.")
            return

        password = simpledialog.askstring(
            "Admin Login", "Enter admin password:", show="*", parent=self.root)
        if password is None:
            return
        if password != self.admin_password:
            messagebox.showerror("Access Denied", "Incorrect admin password.")
            return

        self.is_admin = True
        self.admin_btn.configure(text=self._effective_admin_button_text())
        if hasattr(self, "columns_frame"):
            self.admin_tools_frame.pack(
                fill="x", pady=(0, 4), before=self.columns_frame)
        else:
            self.admin_tools_frame.pack(fill="x", pady=(0, 4))
        self.admin_tools_frame.lift()
        self.admin_tools_frame.update_idletasks()
        self._refresh_intern_summary()
        self._refresh_main_ui_window()

    def _admin_logout(self):
        if not self.is_admin:
            return
        self.is_admin = False
        self.admin_btn.configure(text=self._effective_admin_button_text())
        self.admin_tools_frame.pack_forget()
        self._refresh_intern_summary()
        self._refresh_main_ui_window()
        try:
            self._open_main_ui_window()
            if self.main_ui_window is not None and self.main_ui_window.winfo_exists():
                try:
                    self.main_ui_window.deiconify()
                    self.main_ui_window.state("zoomed")
                    self.main_ui_window.lift()
                    self.main_ui_window.focus_force()
                except Exception:
                    pass
            try:
                self.root.withdraw()
            except Exception:
                pass
        except Exception:
            if self.main_ui_window is not None and self.main_ui_window.winfo_exists():
                self.main_ui_window.deiconify()
                try:
                    self.main_ui_window.state("zoomed")
                except Exception:
                    pass
                self.main_ui_window.lift()
                try:
                    self.root.withdraw()
                except Exception:
                    pass

    def _change_admin_password(self):
        if not self.is_admin:
            messagebox.showwarning(
                "Admin Only", "Log in as admin before changing password.")
            return

        current = simpledialog.askstring(
            "Change Password", "Current password:", show="*", parent=self.root)
        if current is None:
            return
        if current != self.admin_password:
            messagebox.showerror(
                "Access Denied", "Current password is incorrect.")
            return

        new_password = simpledialog.askstring(
            "Change Password", "New password:", show="*", parent=self.root)
        if new_password is None:
            return
        new_password = new_password.strip()
        if len(new_password) < 4:
            messagebox.showerror("Invalid Password",
                                 "Password must be at least 4 characters.")
            return

        confirm = simpledialog.askstring(
            "Change Password", "Confirm new password:", show="*", parent=self.root)
        if confirm is None:
            return
        if new_password != confirm.strip():
            messagebox.showerror(
                "Mismatch", "New password and confirmation do not match.")
            return

        self.admin_password = new_password
        self._save_admin_password()
        messagebox.showinfo("Success", "Admin password updated.")

    def _change_selected_user_password(self):
        if not self._require_admin():
            return
        user = self._active_user()
        if not user or user not in self.users_data:
            messagebox.showerror("No User", "Select a user first.")
            return
        new_password = simpledialog.askstring(
            "Change User Password",
            f"New password for {self._full_name_for_user(user)}:",
            show="*",
            parent=self.root,
        )
        if new_password is None:
            return
        new_password = new_password.strip()
        if len(new_password) < 4:
            messagebox.showerror("Invalid Password",
                                 "Password must be at least 4 characters.")
            return
        confirm = simpledialog.askstring(
            "Change User Password", "Confirm new password:", show="*", parent=self.root)
        if confirm is None:
            return
        if new_password != confirm.strip():
            messagebox.showerror(
                "Mismatch", "Password and confirmation do not match.")
            return
        self.users_data[user]["password"] = _hash_password(new_password)
        self._save_records()
        messagebox.showinfo(
            "Success", f"Password updated for {self._full_name_for_user(user)}.")

    def _edit_selected_user_details(self):
        if not self._require_admin():
            return
        user = self._active_user()
        if not user or user not in self.users_data:
            messagebox.showerror("No User", "Select a user first.")
            return

        data = self.users_data[user]
        dialog = tk.Toplevel(self.root)
        dialog.title("Edit User Details")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(False, False)
        form = ttk.Frame(dialog, padding=12)
        form.pack(fill="both", expand=True)

        vars_map = {
            "first_name": tk.StringVar(value=str(data.get("first_name", ""))),
            "last_name": tk.StringVar(value=str(data.get("last_name", ""))),
            "job_role": tk.StringVar(value=str(data.get("job_role", ""))),
            "required_hours": tk.StringVar(value=str(data.get("required_hours", "500"))),
            "student_id": tk.StringVar(value=str(data.get("student_id", ""))),
            "course": tk.StringVar(value=str(data.get("course", ""))),
            "university": tk.StringVar(value=str(data.get("university", ""))),
        }
        labels = [
            ("First Name", "first_name"),
            ("Last Name", "last_name"),
            ("Job Role", "job_role"),
            ("Required Hours", "required_hours"),
            ("Student ID", "student_id"),
            ("Course", "course"),
            ("University", "university"),
        ]
        for idx, (label, key) in enumerate(labels):
            ttk.Label(form, text=label + ":").grid(row=idx,
                                                   column=0, sticky="w", pady=(0, 6))
            ttk.Entry(form, textvariable=vars_map[key], width=30).grid(row=idx, column=1, sticky="ew", padx=(8, 0),
                                                                       pady=(0, 6))
        form.grid_columnconfigure(1, weight=1)

        def on_save():
            first = vars_map["first_name"].get().strip()
            last = vars_map["last_name"].get().strip()
            required = vars_map["required_hours"].get().strip()
            if not first or not last:
                messagebox.showerror(
                    "Invalid Input", "First Name and Last Name are required.", parent=dialog)
                return
            try:
                req_val = float(required)
                if req_val < 0:
                    raise ValueError
            except ValueError:
                messagebox.showerror(
                    "Invalid Input", "Required Hours must be a non-negative number.", parent=dialog)
                return

            new_key = f"{first} {last}".strip()
            if new_key != user and new_key in self.users_data:
                messagebox.showerror(
                    "Duplicate User", "A user with that full name already exists.", parent=dialog)
                return

            payload = self.users_data[user]
            payload["first_name"] = first
            payload["last_name"] = last
            payload["job_role"] = vars_map["job_role"].get().strip()
            payload["required_hours"] = f"{req_val:.2f}"
            payload["student_id"] = vars_map["student_id"].get().strip()
            payload["course"] = vars_map["course"].get().strip()
            payload["university"] = vars_map["university"].get().strip()

            if new_key != user:
                self.users_data.pop(user, None)
                self.users_data[new_key] = payload
                self.active_user_var.set(new_key)
                self._last_active_user = new_key
            else:
                self.active_user_var.set(user)
                self._last_active_user = user

            self._refresh_user_combo()
            self._load_active_user_into_view()
            self._save_records()
            self._refresh_main_ui_window()
            dialog.destroy()

        actions = ttk.Frame(form)
        actions.grid(row=len(labels), column=0,
                     columnspan=2, sticky="e", pady=(8, 0))
        ttk.Button(actions, text="Cancel",
                   command=dialog.destroy).pack(side="right")
        ttk.Button(actions, text="Save", command=on_save).pack(
            side="right", padx=(0, 8))
        self._center_toplevel(dialog)
        self.root.wait_window(dialog)

    def _require_admin(self):
        if self.is_admin:
            return True
        messagebox.showwarning(
            "Admin Only", "Only admin/head can edit date, time in, and time out.")
        return False

    def _export_active_records_csv(self):
        if not self._require_admin():
            return
        user = self._active_user()
        if user not in self.users_data:
            return
        path = filedialog.asksaveasfilename(
            parent=self.root,
            defaultextension=".csv",
            filetypes=[("CSV", "*.csv")],
            initialfile=f"{user}_attendance.csv",
        )
        if not path:
            return
        with open(path, "w", newline="", encoding="utf-8") as file:
            writer = csv.writer(file)
            writer.writerow(["Date", "Time_In", "Time_Out", "Hours", "Status"])
            for row in self.users_data[user].get("records", []):
                if len(row) >= 5:
                    writer.writerow(row[:5])
                elif len(row) == 4:
                    writer.writerow(
                        [row[0], row[1], row[2], row[3], "On Time"])
        messagebox.showinfo("Export", f"CSV exported:\n{path}")

    def _export_active_records_pdf(self):
        if not self._require_admin():
            return
        if pdf_canvas is None:
            messagebox.showerror(
                "PDF Export", "reportlab is not installed. Install with: pip install reportlab")
            return
        user = self._active_user()
        if user not in self.users_data:
            return
        path = filedialog.asksaveasfilename(
            parent=self.root,
            defaultextension=".pdf",
            filetypes=[("PDF", "*.pdf")],
            initialfile=f"{user}_attendance.pdf",
        )
        if not path:
            return
        records = self.users_data[user].get("records", [])
        total_hours = 0.0
        for row in records:
            try:
                total_hours += float(row[3])
            except Exception:
                continue

        c = pdf_canvas.Canvas(path, pagesize=letter)
        w, h = letter
        y = h - 40
        c.setFont("Helvetica-Bold", 14)
        c.drawString(40, y, f"Attendance Report - {user}")
        y -= 18
        c.setFont("Helvetica-Bold", 10)
        c.drawString(40, y, f"Total Hours: {total_hours:.2f}")
        y -= 14
        c.setFont("Helvetica", 10)
        c.drawString(40, y, "Date")
        c.drawString(140, y, "Time In")
        c.drawString(240, y, "Time Out")
        c.drawString(340, y, "Hours")
        c.drawString(420, y, "Status")
        y -= 15
        for row in records:
            if y < 40:
                c.showPage()
                y = h - 40
            if len(row) >= 5:
                d, tin, tout, hrs, stat = row[:5]
            else:
                d, tin, tout, hrs = row[:4]
                stat = "On Time"
            c.drawString(40, y, str(d))
            c.drawString(140, y, str(tin))
            c.drawString(240, y, str(tout))
            c.drawString(340, y, str(hrs))
            c.drawString(420, y, str(stat))
            y -= 14
        if y < 40:
            c.showPage()
            y = h - 40
        y -= 8
        c.setFont("Helvetica-Bold", 10)
        c.drawString(40, y, f"Total Hours: {total_hours:.2f}")
        c.save()
        messagebox.showinfo("Export", f"PDF exported:\n{path}")

    def _add_user(self):
        if not self._require_admin():
            return
        self._commit_user_from_view(self._active_user())
        profile = self._prompt_new_user_details()
        if profile is None:
            return
        first = profile["first_name"].strip()
        last = profile["last_name"].strip()
        name = f"{first} {last}".strip()
        if name in self.users_data:
            messagebox.showerror("Error", "User already exists.")
            return

        required = profile["required_hours"].strip()
        self.users_data[name] = self._new_user_payload(
            required_hours=required, profile=profile)
        self.active_user_var.set(name)
        self._last_active_user = name
        self._refresh_user_combo()
        self._load_active_user_into_view()
        self._save_records()
        self._refresh_main_ui_window()

    def _delete_user(self):
        if not self._require_admin():
            return
        self._commit_user_from_view(self._active_user())
        active_users = sorted(self._active_user_keys(),
                              key=lambda key: self._full_name_for_user(key).lower())
        if len(active_users) <= 1:
            messagebox.showerror("Error", "At least one user is required.")
            return

        dialog = tk.Toplevel(self.root)
        dialog.title("Delete User Permanently")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(False, False)
        form = ttk.Frame(dialog, padding=12)
        form.pack(fill="both", expand=True)

        user_display_map = {user: self._display_label_for_user(
            user) for user in active_users}
        display_to_user = {display: user for user,
                           display in user_display_map.items()}
        selected_display = tk.StringVar(
            value=user_display_map[active_users[0]])

        ttk.Label(form, text="Select user to delete:").grid(
            row=0, column=0, sticky="w")
        user_combo = ttk.Combobox(
            form,
            textvariable=selected_display,
            values=[user_display_map[user] for user in active_users],
            state="readonly",
            width=35,
        )
        user_combo.grid(row=1, column=0, sticky="ew", pady=(6, 0))
        form.grid_columnconfigure(0, weight=1)

        def on_delete():
            selected = selected_display.get().strip()
            user = display_to_user.get(selected, "")
            if not user:
                messagebox.showerror("Invalid Selection",
                                     "Select a valid user.", parent=dialog)
                return
            if not messagebox.askyesno("Confirm", f"Delete User '{self._full_name_for_user(user)}' and all records?",
                                       parent=dialog):
                return
            self.users_data.pop(user, None)
            remaining = sorted(self._active_user_keys(),
                               key=lambda key: self._full_name_for_user(key).lower())
            self.active_user_var.set(remaining[0] if remaining else "")
            self._last_active_user = self.active_user_var.get()
            self._refresh_user_combo()
            self._load_active_user_into_view()
            self._save_records()
            self._refresh_main_ui_window()
            dialog.destroy()

        actions = ttk.Frame(form)
        actions.grid(row=2, column=0, sticky="e", pady=(10, 0))
        ttk.Button(actions, text="Cancel",
                   command=dialog.destroy).pack(side="right")
        ttk.Button(actions, text="Delete Permanently",
                   command=on_delete).pack(side="right", padx=(0, 8))
        user_combo.focus_set()
        self._center_toplevel(dialog)
        self.root.wait_window(dialog)

    def _prompt_new_user_details(self):
        dialog = tk.Toplevel(self.root)
        dialog.title("Add User")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(False, False)
        form = ttk.Frame(dialog, padding=12)
        form.pack(fill="both", expand=True)
        vars_map = {
            "first_name": tk.StringVar(),
            "last_name": tk.StringVar(),
            "job_role": tk.StringVar(),
            "required_hours": tk.StringVar(value="500"),
            "student_id": tk.StringVar(),
            "course": tk.StringVar(),
            "university": tk.StringVar(),
            "password": tk.StringVar(),
        }
        labels = [
            ("First Name", "first_name"),
            ("Last Name", "last_name"),
            ("Job Role", "job_role"),
            ("Required Hours", "required_hours"),
            ("Student ID", "student_id"),
            ("Course", "course"),
            ("University", "university"),
            ("Password", "password"),
        ]
        for idx, (label, key) in enumerate(labels):
            ttk.Label(form, text=label + ":").grid(row=idx,
                                                   column=0, sticky="w", pady=(0, 6))
            if key == "password":
                ttk.Entry(form, textvariable=vars_map[key], width=30, show="*").grid(row=idx, column=1, sticky="ew",
                                                                                     padx=(8, 0), pady=(0, 6))
            else:
                ttk.Entry(form, textvariable=vars_map[key], width=30).grid(row=idx, column=1, sticky="ew", padx=(8, 0),
                                                                           pady=(0, 6))
        form.grid_columnconfigure(1, weight=1)
        result = {"value": None}

        def on_save():
            first = vars_map["first_name"].get().strip()
            last = vars_map["last_name"].get().strip()
            required = vars_map["required_hours"].get().strip()
            password = vars_map["password"].get().strip()
            if not first or not last:
                messagebox.showerror(
                    "Invalid Input", "First Name and Last Name are required.", parent=dialog)
                return
            if len(password) < 4:
                messagebox.showerror(
                    "Invalid Input", "Password must be at least 4 characters.", parent=dialog)
                return
            try:
                val = float(required)
                if val < 0:
                    raise ValueError
            except ValueError:
                messagebox.showerror(
                    "Invalid Input", "Required Hours must be a non-negative number.", parent=dialog)
                return
            result["value"] = {key: var.get().strip()
                               for key, var in vars_map.items()}
            dialog.destroy()

        actions = ttk.Frame(form)
        actions.grid(row=len(labels), column=0,
                     columnspan=2, sticky="e", pady=(6, 0))
        ttk.Button(actions, text="Cancel",
                   command=dialog.destroy).pack(side="right")
        ttk.Button(actions, text="Add User", command=on_save).pack(
            side="right", padx=(0, 8))
        self._center_toplevel(dialog)
        self.root.wait_window(dialog)
        return result["value"]

    def _soft_delete_selected_intern(self, action_mode="toggle"):
        if not self._require_admin():
            return
        selected = self.summary_tree.selection()
        if not selected:
            messagebox.showwarning("Selection Required",
                                   "Select an intern from Intern's Summary.")
            return
        user = selected[0]
        if user not in self.users_data:
            return
        current = self.users_data[user].get("is_deleted", False)
        if action_mode == "delete":
            if current:
                return
            next_deleted = True
            action = "soft delete"
        elif action_mode == "restore":
            if not current:
                return
            next_deleted = False
            action = "restore"
        else:
            next_deleted = not current
            action = "restore" if current else "soft delete"
        if not messagebox.askyesno("Confirm", f"Do you want to {action} '{self._full_name_for_user(user)}'?"):
            return
        self.users_data[user]["is_deleted"] = next_deleted
        active_users = self._active_user_keys()
        if self._active_user() == user and self.users_data[user]["is_deleted"]:
            if active_users:
                self.active_user_var.set(active_users[0])
            else:
                self.active_user_var.set("")
        self._refresh_user_combo()
        self._load_active_user_into_view()
        self._save_records()
        self._refresh_main_ui_window()

    def _today_text(self):
        return date.today().strftime(DATE_FMT)

    def _safe_filename(self, text: str):
        cleaned = re.sub(r'[^A-Za-z0-9._ -]+', "_", text).strip()
        return cleaned or "user"

    def _generate_qr_for_active_user(self):
        if not self.is_admin:
            messagebox.showwarning(
                "Admin Only", "Log in as admin to generate QR.")
            return
        if qrcode is None:
            messagebox.showerror("QR Generator Unavailable",
                                 "Install qrcode first: pip install qrcode[pil]")
            return

        user = self._active_user()
        if not user:
            messagebox.showerror("No User", "Select a user first.")
            return

        QR_DIR.mkdir(parents=True, exist_ok=True)
        payload = f"user:{user}"
        qr_obj = qrcode.QRCode(
            version=None,
            error_correction=qrcode.constants.ERROR_CORRECT_M,
            box_size=10,
            border=4,
        )
        qr_obj.add_data(payload)
        qr_obj.make(fit=True)
        qr_img = qr_obj.make_image(fill_color="black", back_color="white")
        out_path = QR_DIR / f"{self._safe_filename(user)}.png"
        qr_img.save(out_path)
        messagebox.showinfo(
            "QR Generated", f"Saved QR for {user}:\n{out_path}")

    def _extract_user_from_qr(self, payload: str):
        text = payload.strip()
        if not text:
            return None

        if text.lower().startswith("user:"):
            return text.split(":", 1)[1].strip()

        try:
            parsed = json.loads(text)
            if isinstance(parsed, dict):
                for key in ("user", "username", "name"):
                    value = parsed.get(key)
                    if isinstance(value, str) and value.strip():
                        return value.strip()
        except Exception:
            pass

        return text

    def _scan_qr_attendance(self):
        if cv2 is None:
            messagebox.showerror("QR Scanner Unavailable",
                                 "OpenCV is not installed. Install with: pip install opencv-python")
            self.scan_feedback_var.set("Scanner unavailable")
            return

        cap = cv2.VideoCapture(0)
        if not cap.isOpened():
            messagebox.showerror("Camera Error", "Unable to access camera.")
            self.scan_feedback_var.set("Camera error")
            return

        detector = cv2.QRCodeDetector()
        decoded = ""
        window_name = "Scan QR - Press Q to cancel"
        try:
            while True:
                ok, frame = cap.read()
                if not ok:
                    continue

                data, bbox, _ = detector.detectAndDecode(frame)
                if bbox is not None:
                    pts = bbox.astype(int).reshape(-1, 2)
                    for i in range(len(pts)):
                        x1, y1 = pts[i]
                        x2, y2 = pts[(i + 1) % len(pts)]
                        cv2.line(frame, (x1, y1), (x2, y2), (0, 255, 0), 2)

                cv2.imshow(window_name, frame)
                if data:
                    decoded = data
                    break
                if cv2.waitKey(1) & 0xFF == ord("q"):
                    break
        finally:
            cap.release()
            cv2.destroyAllWindows()

        if not decoded:
            self.scan_feedback_var.set("Scan cancelled")
            return

        user = self._extract_user_from_qr(decoded)
        if not user:
            messagebox.showerror("Invalid QR", "No user found in QR data.")
            self.scan_feedback_var.set("Invalid QR detected")
            return
        if user not in self.users_data:
            messagebox.showerror(
                "Unknown User", f"User '{user}' is not registered.")
            self.scan_feedback_var.set(f"Unknown user: {user}")
            return

        if user != self._active_user():
            self.active_user_var.set(user)
            self._on_user_change()

        today = self._today_text()
        if self._find_open_item_for_date(today):
            self._add_time_out()
            self.scan_feedback_var.set(f"Time-out recorded for {user}")
            return
        if self._find_closed_item_for_date(today):
            messagebox.showinfo(
                "Attendance", f"{user} already completed attendance today.")
            self.scan_feedback_var.set(f"{user} already completed today")
            return
        self._add_time_in()
        self.scan_feedback_var.set(f"QR detected: Time-in recorded for {user}")

    def _parse_display_time_on_date(self, work_date: date, time_text: str) -> datetime:
        parsed_time = datetime.strptime(time_text, TIME_FMT_DISPLAY)
        return datetime.combine(work_date, parsed_time.time())

    def _compute_attendance_status(self, dt_in: datetime, hours: float):
        """
        Business rules (per your requirements):
        - Half Day if < 4 hours
        - Late if time-in after 9:00 AM (even if they complete 8+ hours)
        - Undertime if < 8 hours (and not late / half day)
        - On Time otherwise
        """
        if hours < 4:
            return "Half Day"
        late_cutoff = datetime.strptime(
            LATE_AFTER_TIME, TIME_FMT_DISPLAY).time()
        if dt_in.time() > late_cutoff:
            return "Late"
        if hours < 8:
            return "Undertime"
        return "On Time"

    def _row_tag_for_status(self, status: str):
        s = (status or "").strip().lower()
        if s == "late":
            return "late"
        if s == "undertime":
            return "undertime"
        if s == "half day":
            return "half_day"
        if s == "on time":
            return "on_time"
        if s == "open":
            return "open"
        return "open"

    def _tree_records(self):
        records = []
        for item in self.tree.get_children():
            values = self.tree.item(item, "values")
            if len(values) >= 5:
                records.append(
                    (values[0], values[1], values[2], values[3], values[4]))
            else:
                records.append(
                    (values[0], values[1], values[2], values[3], "On Time"))
        return records

    def _refresh_tree_visuals(self):
        items = self.tree.get_children()
        for index, item in enumerate(items):
            values = self.tree.item(item, "values")
            status = values[4] if len(values) >= 5 else "Open"
            zebra = "even" if (index % 2 == 0) else "odd"
            status_tag = self._row_tag_for_status(status)
            self.tree.item(item, tags=(zebra, status_tag))

        if items:
            self.empty_state_label.place_forget()
        else:
            self.empty_state_label.place(relx=0.5, rely=0.5, anchor="center")
        self._update_record_summary()

    def _append_tree_row(self, values):
        self.tree.insert("", "end", values=values)
        self._refresh_tree_visuals()

    def _resize_tree_columns(self, event=None):
        total_width = self.tree.winfo_width()
        if total_width <= 100:
            return
        date_w = int(total_width * 0.20)
        time_in_w = int(total_width * 0.20)
        time_out_w = int(total_width * 0.20)
        hours_w = int(total_width * 0.16)
        status_w = max(total_width - (date_w + time_in_w +
                       time_out_w + hours_w), 120)
        self.tree.column("date", width=date_w)
        self.tree.column("time_in", width=time_in_w)
        self.tree.column("time_out", width=time_out_w)
        self.tree.column("hours", width=hours_w)
        self.tree.column("status", width=status_w)

    def _pick_date_only(self, initial_date=None, title="Select Date"):
        selected = {"date": None}
        if initial_date:
            try:
                today = datetime.strptime(initial_date, DATE_FMT).date()
            except ValueError:
                today = datetime.now().date()
        else:
            today = datetime.now().date()
        state = {"year": today.year, "month": today.month, "day": today.day}

        picker = tk.Toplevel(self.root)
        picker.title(title)
        picker.transient(self.root)
        picker.grab_set()
        picker.resizable(False, False)
        picker.configure(bg="#ffffff", padx=10, pady=10)

        header = tk.Frame(picker, bg="#ffffff")
        header.grid(row=0, column=0, sticky="ew")
        header.columnconfigure(1, weight=1)
        month_label = tk.Label(header, text="", bg="#ffffff",
                               fg="#1f2a44", font=("Segoe UI", 10, "bold"))
        month_label.grid(row=0, column=1)

        days_frame = tk.Frame(picker, bg="#ffffff")
        days_frame.grid(row=1, column=0, pady=(6, 6))

        def move_month(delta):
            month = state["month"] + delta
            year = state["year"]
            if month < 1:
                month = 12
                year -= 1
            elif month > 12:
                month = 1
                year += 1
            state["month"] = month
            state["year"] = year
            render_calendar()

        tk.Button(header, text="◀", command=lambda: move_month(-1), bg="#ffffff", fg="#1f2a44", bd=0,
                  activebackground="#f1f5f9").grid(row=0, column=0, sticky="w")
        tk.Button(header, text="▶", command=lambda: move_month(1), bg="#ffffff", fg="#1f2a44", bd=0,
                  activebackground="#f1f5f9").grid(row=0, column=2, sticky="e")

        def select_day(day):
            state["day"] = day
            render_calendar()

        def render_calendar():
            month_label.configure(
                text=f"{pycalendar.month_name[state['month']]} {state['year']}")
            for widget in days_frame.winfo_children():
                widget.destroy()

            week_names = ["Su", "Mo", "Tu", "We", "Th", "Fr", "Sa"]
            for col, name in enumerate(week_names):
                tk.Label(days_frame, text=name, bg="#ffffff", fg="#1f2a44", font=("Segoe UI", 9, "bold"), width=3).grid(
                    row=0, column=col, padx=2, pady=2)

            weeks = pycalendar.monthcalendar(state["year"], state["month"])
            for row, week in enumerate(weeks, start=1):
                for col, day in enumerate(week):
                    if day == 0:
                        tk.Label(days_frame, text="", bg="#ffffff", width=3).grid(
                            row=row, column=col, padx=2, pady=2)
                        continue
                    is_selected = day == state["day"]
                    tk.Button(
                        days_frame,
                        text=str(day),
                        width=3,
                        command=lambda d=day: select_day(d),
                        bg="#dbeafe" if is_selected else "#ffffff",
                        fg="#1f2a44",
                        bd=1,
                        activebackground="#bfdbfe",
                        activeforeground="#1f2a44",
                    ).grid(row=row, column=col, padx=2, pady=2)

        def on_select():
            selected["date"] = f"{state['year']:04d}-{state['month']:02d}-{state['day']:02d}"
            picker.destroy()

        def on_today():
            now = datetime.now().date()
            state["year"] = now.year
            state["month"] = now.month
            state["day"] = now.day
            render_calendar()

        def on_clear():
            selected["date"] = ""
            picker.destroy()

        actions = tk.Frame(picker, bg="#ffffff")
        actions.grid(row=2, column=0, sticky="ew")
        actions.columnconfigure(0, weight=1)
        actions.columnconfigure(1, weight=1)
        actions.columnconfigure(2, weight=1)
        tk.Button(actions, text="Clear", command=on_clear, bg="#ffffff", fg="#2563eb", bd=0, activebackground="#eff6ff",
                  activeforeground="#2563eb").grid(row=0, column=0, sticky="w")
        tk.Button(actions, text="Today", command=on_today, bg="#ffffff", fg="#2563eb", bd=0, activebackground="#eff6ff",
                  activeforeground="#2563eb").grid(row=0, column=1)
        ttk.Button(actions, text="Select", command=on_select).grid(
            row=0, column=2, sticky="e")

        render_calendar()
        self._center_toplevel(picker)
        picker.protocol("WM_DELETE_WINDOW", picker.destroy)
        self.root.wait_window(picker)
        return selected["date"]

    def _select_from_date(self):
        picked = self._pick_date_only(
            initial_date=self.from_date_var.get().strip(), title="Select From Date")
        if picked is None:
            return
        self.from_date_var.set(picked)

    def _select_to_date(self):
        picked = self._pick_date_only(
            initial_date=self.to_date_var.get().strip(), title="Select To Date")
        if picked is None:
            return
        self.to_date_var.set(picked)

    def _matches_filters(self, values):
        date_text = self.date_filter_var.get().strip().lower()
        time_in_text = self.time_in_filter_var.get().strip().lower()
        time_out_text = self.time_out_filter_var.get().strip().lower()
        from_text = self.from_date_var.get().strip()
        to_text = self.to_date_var.get().strip()

        if date_text and date_text not in str(values[0]).lower():
            return False
        if time_in_text and time_in_text not in str(values[1]).lower():
            return False
        if time_out_text and time_out_text not in str(values[2]).lower():
            return False

        try:
            row_date = datetime.strptime(values[0], DATE_FMT).date()
        except Exception:
            return False

        if from_text:
            try:
                from_date = datetime.strptime(from_text, DATE_FMT).date()
                if row_date < from_date:
                    return False
            except ValueError:
                pass
        if to_text:
            try:
                to_date = datetime.strptime(to_text, DATE_FMT).date()
                if row_date > to_date:
                    return False
            except ValueError:
                pass
        return True

    def _apply_record_filters(self):
        # Reload full data, then show filtered view without mutating stored records.
        self._filters_active = True
        self._load_active_user_into_view(apply_filters=True)

    def _reset_record_filters(self):
        self.date_filter_var.set("")
        self.time_in_filter_var.set("")
        self.time_out_filter_var.set("")
        self.from_date_var.set("")
        self.to_date_var.set("")
        self._filters_active = False
        self._load_active_user_into_view(apply_filters=False)

    def _update_record_summary(self):
        rows = [self.tree.item(item, "values")
                for item in self.tree.get_children()]
        closed = [r for r in rows if len(
            r) >= 5 and r[3] not in ("", None, "")]
        total_days = len(closed)

        avg_hours = 0.0
        hours_values = []
        for r in closed:
            try:
                hours_values.append(float(r[3]))
            except (TypeError, ValueError):
                continue
        if hours_values:
            avg_hours = sum(hours_values) / len(hours_values)

        half_day_count = sum(1 for r in closed if "Half Day" in str(r[4]))
        late_count = sum(1 for r in closed if "Late" in str(r[4]))
        self.summary_var.set(
            f"Total Days: {total_days} | Avg Hours: {avg_hours:.2f} | "
            f"Late Count: {late_count} | Half Day Count: {half_day_count}"
        )

    def _find_open_item_for_date(self, date_text: str):
        for item in reversed(self.tree.get_children()):
            values = self.tree.item(item, "values")
            if len(values) < 3:
                continue
            if values[0] == date_text and values[2] == "":
                return item
        return None

    def _find_closed_item_for_date(self, date_text: str):
        for item in reversed(self.tree.get_children()):
            values = self.tree.item(item, "values")
            if len(values) < 3:
                continue
            if values[0] == date_text and values[2] != "":
                return item
        return None

    def _update_time_button_states(self):
        return

    def _heartbeat(self):
        self.date_var.set(self._today_text())
        now = datetime.now()
        self.overview_date_var.set(f"Date: {now.strftime('%b %d, %Y')}")
        self.overview_time_var.set(f"Time: {now.strftime('%I:%M %p')}")
        self._update_time_button_states()
        self.root.after(30000, self._heartbeat)

    def _add_time_in(self):
        if not self._active_user():
            messagebox.showerror("Error", "Select a user first.")
            return

        dt_in = datetime.now() - timedelta(minutes=TIME_IN_GRACE_MINUTES)
        work_date = dt_in.date()
        date_text = work_date.strftime(DATE_FMT)

        if self._find_open_item_for_date(date_text):
            messagebox.showwarning(
                "Already Timed In", "Already timed in today.")
            self._update_time_button_states()
            return
        if self._find_closed_item_for_date(date_text):
            messagebox.showwarning("Attendance Complete",
                                   "Already timed out today.")
            self._update_time_button_states()
            return

        self._append_tree_row(
            (
                date_text,
                dt_in.strftime(TIME_FMT_DISPLAY),
                "",
                "",
                "Open",
            )
        )
        self.date_var.set(date_text)
        self.time_in_var.set(dt_in.strftime(TIME_FMT_DISPLAY))
        self.time_out_var.set("")
        self._recompute_totals()
        self._update_time_button_states()

    def _add_time_out(self):
        if not self._active_user():
            messagebox.showerror("Error", "Select a user first.")
            return

        dt_out = datetime.now()
        if dt_out.hour < SIGN_OUT_ALLOWED_FROM_HOUR:
            messagebox.showwarning(
                "Too Early", "You can only time out at 5:00 PM onwards.")
            self._update_time_button_states()
            return

        work_date = dt_out.date()
        date_text = work_date.strftime(DATE_FMT)

        if self._find_closed_item_for_date(date_text) and not self._find_open_item_for_date(date_text):
            messagebox.showwarning("Already Timed Out",
                                   "Already timed out today.")
            self._update_time_button_states()
            return

        item = self._find_open_item_for_date(date_text)
        if not item:
            messagebox.showwarning(
                "Missing Time In", "You need to time in first.")
            self._update_time_button_states()
            return

        values = self.tree.item(item, "values")
        dt_in = self._parse_display_time_on_date(work_date, values[1])
        if dt_out <= dt_in:
            dt_out = dt_out + timedelta(days=1)

        hours = (dt_out - dt_in).total_seconds() / 3600.0
        status = self._compute_attendance_status(dt_in, hours)
        self.tree.item(
            item,
            values=(
                values[0],
                values[1],
                dt_out.strftime(TIME_FMT_DISPLAY),
                f"{hours:.2f}",
                status,
            ),
        )
        self.date_var.set(date_text)
        self.time_out_var.set(dt_out.strftime(TIME_FMT_DISPLAY))
        self._recompute_totals()
        self._update_time_button_states()

    def _show_context_menu(self, event):
        item = self.tree.identify_row(event.y)
        if item:
            self.tree.selection_set(item)
            self.tree.focus(item)
        else:
            self.tree.selection_remove(self.tree.selection())
        self.context_menu.tk_popup(event.x_root, event.y_root)

    # (Admin date+time picker and context actions kept mostly as-is)
    def _pick_date_for_admin_entry(self, initial_date=None, initial_time_in="08:30 AM", initial_time_out="05:30 PM",
                                   title="Add Entry"):
        selected = {"date": None, "time_in": None, "time_out": None}
        if initial_date:
            try:
                today = datetime.strptime(initial_date, DATE_FMT).date()
            except ValueError:
                today = datetime.now().date()
        else:
            today = datetime.now().date()
        state = {"year": today.year, "month": today.month, "day": today.day}

        picker = tk.Toplevel(self.root)
        picker.title(title)
        picker.transient(self.root)
        picker.grab_set()
        picker.resizable(False, False)
        picker.configure(bg="#ffffff", padx=10, pady=10)

        header = tk.Frame(picker, bg="#ffffff")
        header.grid(row=0, column=0, sticky="ew")
        header.columnconfigure(1, weight=1)
        month_label = tk.Label(header, text="", bg="#ffffff",
                               fg="#1f2a44", font=("Segoe UI", 10, "bold"))
        month_label.grid(row=0, column=1)

        days_frame = tk.Frame(picker, bg="#ffffff")
        days_frame.grid(row=1, column=0, pady=(6, 6))

        def move_month(delta):
            month = state["month"] + delta
            year = state["year"]
            if month < 1:
                month = 12
                year -= 1
            elif month > 12:
                month = 1
                year += 1
            state["month"] = month
            state["year"] = year
            render_calendar()

        tk.Button(header, text="◀", command=lambda: move_month(-1), bg="#ffffff", fg="#1f2a44", bd=0,
                  activebackground="#f1f5f9").grid(row=0, column=0, sticky="w")
        tk.Button(header, text="▶", command=lambda: move_month(1), bg="#ffffff", fg="#1f2a44", bd=0,
                  activebackground="#f1f5f9").grid(row=0, column=2, sticky="e")

        def select_day(day):
            state["day"] = day
            render_calendar()

        def render_calendar():
            month_label.configure(
                text=f"{pycalendar.month_name[state['month']]} {state['year']}")
            for widget in days_frame.winfo_children():
                widget.destroy()

            week_names = ["Su", "Mo", "Tu", "We", "Th", "Fr", "Sa"]
            for col, name in enumerate(week_names):
                tk.Label(days_frame, text=name, bg="#ffffff", fg="#1f2a44", font=("Segoe UI", 9, "bold"), width=3).grid(
                    row=0, column=col, padx=2, pady=2)

            weeks = pycalendar.monthcalendar(state["year"], state["month"])
            for row, week in enumerate(weeks, start=1):
                for col, day in enumerate(week):
                    if day == 0:
                        tk.Label(days_frame, text="", bg="#ffffff", width=3).grid(
                            row=row, column=col, padx=2, pady=2)
                        continue
                    is_selected = day == state["day"]
                    tk.Button(
                        days_frame,
                        text=str(day),
                        width=3,
                        command=lambda d=day: select_day(d),
                        bg="#dbeafe" if is_selected else "#ffffff",
                        fg="#1f2a44",
                        bd=1,
                        activebackground="#bfdbfe",
                        activeforeground="#1f2a44",
                    ).grid(row=row, column=col, padx=2, pady=2)

        def on_ok():
            time_in_text = time_in_var.get().strip().upper()
            time_out_text = time_out_var.get().strip().upper()
            try:
                datetime.strptime(time_in_text, TIME_FMT_DISPLAY)
                datetime.strptime(time_out_text, TIME_FMT_DISPLAY)
            except ValueError:
                messagebox.showerror(
                    "Invalid Input", "Use valid time format: HH:MM AM/PM.", parent=picker)
                return

            selected["date"] = f"{state['year']:04d}-{state['month']:02d}-{state['day']:02d}"
            selected["time_in"] = time_in_text
            selected["time_out"] = time_out_text
            picker.destroy()

        def on_cancel():
            picker.destroy()

        def on_today():
            now = datetime.now().date()
            state["year"] = now.year
            state["month"] = now.month
            state["day"] = now.day
            render_calendar()

        def on_clear():
            selected["date"] = None
            picker.destroy()

        time_section = tk.Frame(picker, bg="#ffffff")
        time_section.grid(row=2, column=0, sticky="ew", pady=(4, 8))
        tk.Label(time_section, text="Time In (HH:MM AM/PM)", bg="#ffffff", fg="#1f2a44", font=("Segoe UI", 9)).grid(
            row=0, column=0, sticky="w")
        tk.Label(time_section, text="Time Out (HH:MM AM/PM)", bg="#ffffff", fg="#1f2a44", font=("Segoe UI", 9)).grid(
            row=0, column=1, padx=(12, 0), sticky="w")
        time_in_var = tk.StringVar(value=(initial_time_in or "08:30 AM"))
        time_out_var = tk.StringVar(value=(initial_time_out or "05:30 PM"))
        time_in_entry = ttk.Entry(
            time_section, textvariable=time_in_var, width=14)
        time_in_entry.grid(row=1, column=0, pady=(4, 0), sticky="w")
        time_out_entry = ttk.Entry(
            time_section, textvariable=time_out_var, width=14)
        time_out_entry.grid(row=1, column=1, padx=(12, 0),
                            pady=(4, 0), sticky="w")

        actions = tk.Frame(picker, bg="#ffffff")
        actions.grid(row=3, column=0, sticky="ew")
        actions.columnconfigure(0, weight=1)
        actions.columnconfigure(1, weight=1)
        actions.columnconfigure(2, weight=1)
        tk.Button(actions, text="Clear", command=on_clear, bg="#ffffff", fg="#2563eb", bd=0, activebackground="#eff6ff",
                  activeforeground="#2563eb").grid(row=0, column=0, sticky="w")
        tk.Button(actions, text="Today", command=on_today, bg="#ffffff", fg="#2563eb", bd=0, activebackground="#eff6ff",
                  activeforeground="#2563eb").grid(row=0, column=1)
        ttk.Button(actions, text="Select", command=on_ok).grid(
            row=0, column=2, sticky="e")

        render_calendar()
        time_in_entry.focus_set()
        self._center_toplevel(picker)
        picker.protocol("WM_DELETE_WINDOW", on_cancel)
        self.root.wait_window(picker)
        if selected["date"] is None:
            return None
        return selected["date"], selected["time_in"], selected["time_out"]

    def _add_entry_from_context(self):
        if not self._require_admin():
            return

        picked = self._pick_date_for_admin_entry()
        if picked is None:
            return
        date_text, time_in_text, time_out_text = picked

        date_text = date_text.strip()
        time_in_text = time_in_text.strip().upper()
        time_out_text = time_out_text.strip().upper()
        if not date_text or not time_in_text or not time_out_text:
            messagebox.showerror(
                "Invalid Input", "Date, Time In, and Time Out are required.")
            return

        try:
            work_date = datetime.strptime(date_text, DATE_FMT).date()
            dt_in = self._parse_display_time_on_date(work_date, time_in_text)
            dt_out = self._parse_display_time_on_date(work_date, time_out_text)
        except ValueError:
            messagebox.showerror(
                "Invalid Input", "Use valid formats: YYYY-MM-DD and HH:MM AM/PM.")
            return

        if dt_out <= dt_in:
            dt_out = dt_out + timedelta(days=1)

        hours = (dt_out - dt_in).total_seconds() / 3600.0
        status = self._compute_attendance_status(dt_in, hours)
        self._append_tree_row(
            (
                work_date.strftime(DATE_FMT),
                dt_in.strftime(TIME_FMT_DISPLAY),
                dt_out.strftime(TIME_FMT_DISPLAY),
                f"{hours:.2f}",
                status,
            )
        )
        self._recompute_totals()

    def _add_date_only_from_context(self):
        if not self._require_admin():
            return
        date_text = simpledialog.askstring("Add Date", "Date (YYYY-MM-DD):", initialvalue=self._today_text(),
                                           parent=self.root)
        if date_text is None:
            return
        date_text = date_text.strip()
        try:
            work_date = datetime.strptime(date_text, DATE_FMT).date()
        except ValueError:
            messagebox.showerror(
                "Invalid Input", "Use valid date format: YYYY-MM-DD.")
            return
        self._append_tree_row(
            (work_date.strftime(DATE_FMT), "", "", "", "Open"))
        self._recompute_totals()

    def _add_time_in_only_from_context(self):
        if not self._require_admin():
            return
        selected_items = self.tree.selection()
        if not selected_items:
            messagebox.showerror(
                "No Selection", "Select an entry to add Time In.")
            return
        item = selected_items[0]
        values = self.tree.item(item, "values")
        if len(values) < 5:
            return
        if not values[0]:
            messagebox.showerror("Invalid Entry", "Set date first.")
            return
        time_in_text = simpledialog.askstring(
            "Add Time In", "Time In (HH:MM AM/PM):", parent=self.root)
        if time_in_text is None:
            return
        time_in_text = time_in_text.strip().upper()
        try:
            work_date = datetime.strptime(values[0], DATE_FMT).date()
            dt_in = self._parse_display_time_on_date(work_date, time_in_text)
        except ValueError:
            messagebox.showerror(
                "Invalid Input", "Use valid time format: HH:MM AM/PM.")
            return
        self.tree.item(item, values=(values[0], dt_in.strftime(
            TIME_FMT_DISPLAY), values[2], values[3], values[4]))
        self._recompute_item_hours(item)
        self._refresh_tree_visuals()
        self._recompute_totals()

    def _add_time_out_only_from_context(self):
        if not self._require_admin():
            return
        selected_items = self.tree.selection()
        if not selected_items:
            messagebox.showerror(
                "No Selection", "Select an entry to add Time Out.")
            return
        item = selected_items[0]
        values = self.tree.item(item, "values")
        if len(values) < 5:
            return
        if not values[0] or not values[1]:
            messagebox.showerror(
                "Invalid Entry", "Set date and Time In first.")
            return
        time_out_text = simpledialog.askstring(
            "Add Time Out", "Time Out (HH:MM AM/PM):", parent=self.root)
        if time_out_text is None:
            return
        time_out_text = time_out_text.strip().upper()
        try:
            work_date = datetime.strptime(values[0], DATE_FMT).date()
            dt_out = self._parse_display_time_on_date(work_date, time_out_text)
        except ValueError:
            messagebox.showerror(
                "Invalid Input", "Use valid time format: HH:MM AM/PM.")
            return
        self.tree.item(item, values=(values[0], values[1], dt_out.strftime(
            TIME_FMT_DISPLAY), values[3], values[4]))
        self._recompute_item_hours(item)
        self._refresh_tree_visuals()
        self._recompute_totals()

    def _recompute_item_hours(self, item):
        values = self.tree.item(item, "values")
        if len(values) < 5:
            return
        date_text, time_in_text, time_out_text, _, _status = values
        if not date_text or not time_in_text:
            return
        try:
            work_date = datetime.strptime(date_text, DATE_FMT).date()
            dt_in = self._parse_display_time_on_date(work_date, time_in_text)
            if time_out_text:
                dt_out = self._parse_display_time_on_date(
                    work_date, time_out_text)
                if dt_out <= dt_in:
                    dt_out = dt_out + timedelta(days=1)
                hours_text = f"{(dt_out - dt_in).total_seconds() / 3600.0:.2f}"
                status = self._compute_attendance_status(
                    dt_in, float(hours_text))
            else:
                hours_text = ""
                status = "Open"
        except ValueError:
            return
        self.tree.item(item, values=(date_text, time_in_text,
                       time_out_text, hours_text, status))

    def _edit_selected_date(self):
        if not self._require_admin():
            return

        selected = self.tree.selection()
        if not selected:
            return

        item = selected[0]
        values = self.tree.item(item, "values")
        if len(values) < 5:
            return

        date_text = simpledialog.askstring(
            "Edit Date", "Date (YYYY-MM-DD):", initialvalue=values[0], parent=self.root)
        if date_text is None:
            return

        date_text = date_text.strip()
        try:
            datetime.strptime(date_text, DATE_FMT)
        except ValueError:
            messagebox.showerror(
                "Invalid Input", "Use valid date format: YYYY-MM-DD.")
            return

        self.tree.item(item, values=(
            date_text, values[1], values[2], values[3], values[4]))
        self._recompute_item_hours(item)
        self._recompute_totals()
        self._update_time_button_states()

    def _edit_selected_time_in(self):
        if not self._require_admin():
            return

        selected = self.tree.selection()
        if not selected:
            return
        item = selected[0]
        values = self.tree.item(item, "values")
        if len(values) < 5:
            return
        time_in_text = simpledialog.askstring("Edit Time In", "Time In (HH:MM AM/PM):", initialvalue=values[1],
                                              parent=self.root)
        if time_in_text is None:
            return
        time_in_text = time_in_text.strip().upper()
        try:
            work_date = datetime.strptime(values[0], DATE_FMT).date()
            dt_in = self._parse_display_time_on_date(work_date, time_in_text)
        except ValueError:
            messagebox.showerror(
                "Invalid Input", "Use valid time format: HH:MM AM/PM.")
            return
        self.tree.item(item, values=(values[0], dt_in.strftime(
            TIME_FMT_DISPLAY), values[2], values[3], values[4]))
        self._recompute_item_hours(item)
        self._recompute_totals()
        self._update_time_button_states()

    def _edit_selected_time_out(self):
        if not self._require_admin():
            return

        selected = self.tree.selection()
        if not selected:
            return
        item = selected[0]
        values = self.tree.item(item, "values")
        if len(values) < 5:
            return
        time_out_text = simpledialog.askstring(
            "Edit Time Out",
            "Time Out (HH:MM AM/PM). Leave blank for open entry:",
            initialvalue=values[2],
            parent=self.root,
        )
        if time_out_text is None:
            return
        time_out_text = time_out_text.strip().upper()
        if time_out_text:
            try:
                work_date = datetime.strptime(values[0], DATE_FMT).date()
                dt_out = self._parse_display_time_on_date(
                    work_date, time_out_text)
                time_out_text = dt_out.strftime(TIME_FMT_DISPLAY)
            except ValueError:
                messagebox.showerror(
                    "Invalid Input", "Use valid time format: HH:MM AM/PM.")
                return
        self.tree.item(item, values=(
            values[0], values[1], time_out_text, values[3], values[4]))
        self._recompute_item_hours(item)
        self._recompute_totals()
        self._update_time_button_states()

    def _delete_selected(self):
        if not self._require_admin():
            return

        selected = self.tree.selection()
        if not selected:
            return
        for item in selected:
            self.tree.delete(item)
        self._refresh_tree_visuals()
        self._recompute_totals()

    def _load_active_user_into_view(self, apply_filters=None):
        user = self._active_user()
        if apply_filters is None:
            apply_filters = self._filters_active

        if not user or user not in self.users_data:
            self._refresh_profile_details()
            self._refresh_intern_summary()
            self._refresh_user_profile_window()
            self._refresh_main_ui_window()
            return

        for item in self.tree.get_children():
            self.tree.delete(item)

        user_data = self.users_data[user]
        self.required_hours_var.set(user_data.get("required_hours", "500"))

        # Insert rows (optionally filtered)
        for row in user_data.get("records", []):
            if len(row) >= 5:
                values = row[:5]
            elif len(row) == 4:
                # normalize old record
                try:
                    work_date = datetime.strptime(row[0], DATE_FMT).date()
                    dt_in = self._parse_display_time_on_date(work_date, row[1])
                    hours_val = float(row[3]) if row[3] else 0.0
                    status = "Open" if not row[2] else self._compute_attendance_status(
                        dt_in, hours_val)
                except Exception:
                    status = "On Time"
                values = (row[0], row[1], row[2], row[3], status)
            else:
                continue

            if apply_filters and not self._matches_filters(values):
                continue
            self.tree.insert("", "end", values=values)

        self._refresh_tree_visuals()
        self._recompute_totals()
        self._update_time_button_states()
        self._refresh_profile_details()
        self._refresh_intern_summary()
        self._refresh_user_profile_window()
        self._refresh_main_ui_window()

    def _recompute_totals(self):
        # Compute from stored records (NOT the filtered tree view) so totals stay correct.
        user = self._active_user()
        completed = 0.0
        if user in self.users_data:
            for rec in self.users_data[user].get("records", []):
                try:
                    if len(rec) > 3 and rec[3] not in ("", None):
                        completed += float(rec[3])
                except Exception:
                    continue

        try:
            required = float(self.required_hours_var.get().strip())
            if required < 0:
                raise ValueError
        except ValueError:
            messagebox.showerror("Invalid Required Hours",
                                 "Required hours must be a non-negative number.")
            return

        remaining = max(required - completed, 0.0)
        self.required_hours_display_var.set(f"{required:.2f}")
        self.completed_hours_var.set(f"{completed:.2f} hours")
        self.remaining_hours_var.set(f"{remaining:.2f} hours")
        self.profile_required_hours_var.set(f"{required:.2f}")
        progress = 0.0 if required == 0 else min(
            (completed / required) * 100.0, 100.0)
        try:
            self._animate_progress(progress)
        except Exception:
            self.progress_value_var.set(progress)
            self.progress_percent_var.set(f"{progress:.2f}%")
        self.overview_completion_var.set(f"Completion: {progress:.2f}%")
        if progress < 20:
            self.status_var.set("Status: Internship Just Started")
            self.status_label.configure(fg=DANGER)
        elif progress < 40:
            self.status_var.set("Status: Progress Ongoing")
            self.status_label.configure(fg=WARNING)
        elif progress < 70:
            self.status_var.set("Status: Midway Completed")
            self.status_label.configure(fg=ACCENT)
        elif progress < 100:
            self.status_var.set("Status: Near Completion")
            self.status_label.configure(fg=SUCCESS)
        else:
            self.status_var.set("Status: Internship Completed")
            self.status_label.configure(fg=SUCCESS)

        if user in self.users_data:
            self.users_data[user]["required_hours"] = self.required_hours_var.get(
            ).strip()

            # Only save records from the tree if the tree is showing full data (no filters).
            if not self._filters_active:
                self.users_data[user]["records"] = self._tree_records()
                self._save_records()

        self._refresh_intern_summary()
        self._refresh_user_profile_window()
        self._refresh_main_ui_window()
        self._update_time_button_states()

    def _save_records(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "records"
        sheet.append([
            "Type", "User", "Required_Hours", "Date", "Time_In", "Time_Out", "Hours", "Status",
            "First_Name", "Last_Name", "Job_Role", "Student_ID", "Course", "University", "Password", "Is_Deleted"
        ])

        for user, data in self.users_data.items():
            sheet.append([
                "meta", user, data.get(
                    "required_hours", "500"), "", "", "", "", "",
                data.get("first_name", ""), data.get(
                    "last_name", ""), data.get("job_role", ""),
                data.get("student_id", ""), data.get(
                    "course", ""), data.get("university", ""),
                data.get("password", ""), "1" if data.get(
                    "is_deleted") else "0",
            ])
            for row in data.get("records", []):
                if len(row) >= 5:
                    sheet.append(
                        ["record", user, "", row[0], row[1], row[2], row[3], row[4], "", "", "", "", "", "", "", ""])
                else:
                    sheet.append(
                        ["record", user, "", row[0], row[1], row[2], row[3], "On Time", "", "", "", "", "", "", "", ""])

        workbook.save(XLSX_FILE)

        with CSV_FILE.open("w", newline="", encoding="utf-8") as file:
            writer = csv.writer(file)
            writer.writerow([
                "Type", "User", "Required_Hours", "Date", "Time_In", "Time_Out", "Hours", "Status",
                "First_Name", "Last_Name", "Job_Role", "Student_ID", "Course", "University", "Password", "Is_Deleted"
            ])
            for user, data in self.users_data.items():
                writer.writerow([
                    "meta", user, data.get(
                        "required_hours", "500"), "", "", "", "", "",
                    data.get("first_name", ""), data.get(
                        "last_name", ""), data.get("job_role", ""),
                    data.get("student_id", ""), data.get(
                        "course", ""), data.get("university", ""),
                    data.get("password", ""), "1" if data.get(
                        "is_deleted") else "0",
                ])
                for row in data.get("records", []):
                    if len(row) >= 5:
                        writer.writerow(
                            ["record", user, "", row[0], row[1], row[2], row[3], row[4], "", "", "", "", "", "", "", ""])
                    else:
                        writer.writerow(
                            ["record", user, "", row[0], row[1], row[2], row[3], "On Time", "", "", "", "", "", "", "",
                             ""])

    def _on_close(self):
        self._commit_user_from_view(self._active_user())
        self._save_records()
        self.root.destroy()

    def _load_records(self):
        rows = []
        if XLSX_FILE.exists():
            try:
                workbook = load_workbook(XLSX_FILE)
                sheet = workbook.active
                rows = [
                    ["" if cell is None else str(cell) for cell in row]
                    for row in sheet.iter_rows(values_only=True)
                ]
            except Exception:
                rows = []
        elif CSV_FILE.exists():
            try:
                with CSV_FILE.open("r", newline="", encoding="utf-8") as file:
                    rows = list(csv.reader(file))
            except Exception:
                rows = []

        if not rows:
            return

        # New multi-user format
        if rows[0] and len(rows[0]) >= 7 and rows[0][0].strip().lower() in ("kind", "type"):
            for row in rows[1:]:
                if len(row) < 7:
                    continue
                kind, user, required, d, ti, to, hrs = row[:7]
                status = row[7] if len(row) > 7 else ""
                first_name = row[8] if len(row) > 8 else ""
                last_name = row[9] if len(row) > 9 else ""
                job_role = row[10] if len(row) > 10 else ""
                student_id = row[11] if len(row) > 11 else ""
                course = row[12] if len(row) > 12 else ""
                university = row[13] if len(row) > 13 else ""
                password = row[14] if len(row) > 14 else ""
                is_deleted = row[15] if len(row) > 15 else "0"
                if not user:
                    continue
                if user not in self.users_data:
                    self.users_data[user] = self._new_user_payload(
                        required_hours="500")
                if kind == "meta":
                    self.users_data[user]["required_hours"] = required or "500"
                    self.users_data[user]["first_name"] = first_name.strip()
                    self.users_data[user]["last_name"] = last_name.strip()
                    self.users_data[user]["job_role"] = job_role.strip()
                    self.users_data[user]["student_id"] = student_id.strip()
                    self.users_data[user]["course"] = course.strip()
                    self.users_data[user]["university"] = university.strip()
                    self.users_data[user]["password"] = password.strip()
                    self.users_data[user]["is_deleted"] = str(
                        is_deleted).strip().lower() in ("1", "true", "yes")
                elif kind == "record":
                    self.users_data[user]["records"].append(
                        (d, ti, to, hrs, status or "On Time"))
            self.users_data = {k: self._normalize_user_payload(
                k, v) for k, v in self.users_data.items()}
            return

        # Backward compatibility with old single-user format.
        default_required = "500"
        start_index = 0
        if rows[0] and rows[0][0] == "required_hours" and len(rows[0]) > 1:
            default_required = rows[0][1]
            start_index = 1

        if len(rows) > start_index and rows[start_index] == ["date", "time_in", "time_out", "hours"]:
            start_index += 1

        records = []
        for row in rows[start_index:]:
            if len(row) < 4:
                continue
            status = row[4] if len(row) > 4 else "On Time"
            records.append((row[0], row[1], row[2], row[3], status))

        self.users_data["Imported User"] = self._new_user_payload(required_hours=default_required,
                                                                  profile={"first_name": "Imported",
                                                                           "last_name": "User"})
        self.users_data["Imported User"]["records"] = records


def main():
    root = tk.Tk()
    app = OJTTrackerApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
